"""Consolidate VFO_Keyer_Order.xlsx — group rows by (Value, Package),
sum Qty / On Hand / Ordered / Received, concatenate ref designators.
Preserves every user-entered value (check-off counts, prices, P/N
corrections, notes are joined). Backs up before modifying."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil
import os
import re
from collections import OrderedDict

XLSX = 'Documentation/VFO_Keyer_Order.xlsx'
BAK = 'Documentation/VFO_Keyer_Order.xlsx.preconsolidation.bak'

if os.path.exists(BAK):
    os.remove(BAK)
shutil.copy(XLSX, BAK)
print(f'Backup: {BAK}')

wb = openpyxl.load_workbook(XLSX)
ws = wb['VFO + Keyer Order']

# Capture all current rows
all_rows = []
for r in range(2, ws.max_row + 1):
    cells = [ws.cell(row=r, column=c).value for c in range(1, 16)]
    all_rows.append(cells)

# Identify section headers vs data rows vs total row
def is_section_header(c):
    return (c[0] is not None
            and (c[1] is None or c[1] == '')
            and isinstance(c[2], str) and c[2].startswith('---'))

def is_total_row(c):
    return any(isinstance(x, str) and x == 'TOTAL:' for x in c)

# Build a list of (section_name, section_header_cells, [data_rows])
sections = []  # in order
current = None  # (section, header, [rows])
for cells in all_rows:
    if is_total_row(cells):
        continue
    if is_section_header(cells):
        if current:
            sections.append(current)
        current = (cells[0], cells, [])
    else:
        if cells[1]:  # has a Ref → real data row
            if current is None:
                # Unsectioned data row at the very start — make a default section
                current = (cells[0] or 'UNCAT',
                           [cells[0] or 'UNCAT', '', '--- (uncategorized) ---', '',
                            None, '', '', '', '', '', '', '', None, None, ''],
                           [])
            current[2].append(cells)
if current:
    sections.append(current)

# Helpers
def norm(s):
    """Normalize for grouping: lowercase, collapse whitespace, strip
    parenthetical content (notes like '(match w/ RLb)' or '(or 240 if
    surplus)' don't change what's being ordered, so they shouldn't
    prevent merging)."""
    if s is None: return ''
    s = str(s)
    s = re.sub(r'\s*\([^)]*\)', '', s)   # strip "(...)" annotations
    s = re.sub(r'\s+', ' ', s).strip()
    return s.lower()

def sum_numeric(items):
    s = 0.0
    any_present = False
    for x in items:
        if x is None or x == '':
            continue
        try:
            n = float(x)
            s += n
            any_present = True
        except (ValueError, TypeError):
            pass
    if not any_present:
        return None
    return int(s) if s == int(s) else s

def first_non_empty(items):
    for x in items:
        if x is not None and x != '' and str(x).strip():
            return x
    return None

def join_unique(items, sep=', '):
    seen = []
    for x in items:
        if x is None or x == '':
            continue
        s = str(x).strip()
        if s and s not in seen:
            seen.append(s)
    return sep.join(seen)

# Consolidate within each section
consolidated = []
orig_count = 0
new_count = 0
for section_name, header, data_rows in sections:
    orig_count += len(data_rows)
    # Group by (normalized value, normalized package)
    groups = OrderedDict()
    for cells in data_rows:
        key = (norm(cells[2]), norm(cells[3]))
        groups.setdefault(key, []).append(cells)

    out_rows = []
    for key, group in groups.items():
        refs = join_unique([c[1] for c in group])
        value = first_non_empty([c[2] for c in group])
        package = first_non_empty([c[3] for c in group])
        qty = sum_numeric([c[4] for c in group])
        on_hand = sum_numeric([c[5] for c in group])
        ordered = sum_numeric([c[6] for c in group])
        received = sum_numeric([c[7] for c in group])
        ordered_from = first_non_empty([c[8] for c in group])
        mfr = first_non_empty([c[9] for c in group])
        mouser = first_non_empty([c[10] for c in group])
        digikey = first_non_empty([c[11] for c in group])
        unit = first_non_empty([c[12] for c in group])
        # col 13 = Ext $ formula → regenerate per row
        notes = join_unique([c[14] for c in group], sep='; ')

        out_rows.append([
            section_name, refs, value, package, qty,
            on_hand, ordered, received, ordered_from,
            mfr, mouser, digikey, unit, None, notes
        ])
    new_count += len(out_rows)
    consolidated.append((section_name, header, out_rows))

# Clear all rows except header (row 1)
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

# Style definitions (match generator)
SECTION_COLOR = {
    'VFO': 'D9E1F2', 'KEYER': 'E2EFDA', 'POWER': 'FFF2CC',
    'CONN': 'FCE4D6', 'TEST': 'EDEDED',
}
SECTION_FILL = 'D9E1F2'
CHECK_FILL = 'FFF2CC'
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

# Write consolidated data
out_row = 2
for section_name, header, data_rows in consolidated:
    # Section header row
    for col in range(1, 16):
        c = ws.cell(row=out_row, column=col, value=header[col-1])
        c.fill = PatternFill('solid', fgColor=SECTION_FILL)
        if col == 1:
            c.font = Font(bold=True, size=11)
        elif col == 3:
            c.font = Font(bold=True, italic=True)
    out_row += 1

    # Data rows
    for cells in data_rows:
        fill = PatternFill('solid', fgColor=SECTION_COLOR.get(section_name, 'FFFFFF'))
        for col, val in enumerate(cells, start=1):
            c = ws.cell(row=out_row, column=col, value=val)
            c.border = border
            if col in (5, 6, 7, 8, 9, 13, 14):
                c.alignment = center
            if col in (6, 7, 8, 9):
                c.fill = PatternFill('solid', fgColor=CHECK_FILL)
            else:
                c.fill = fill
        if cells[4] and cells[12]:
            ws.cell(row=out_row, column=14).value = f'=E{out_row}*M{out_row}'
            ws.cell(row=out_row, column=14).number_format = '"$"#,##0.00'
        if cells[12]:
            ws.cell(row=out_row, column=13).number_format = '"$"#,##0.00'
        out_row += 1

# Total row
out_row += 1
tlc = ws.cell(row=out_row, column=12, value='TOTAL:')
tlc.font = Font(bold=True)
tlc.alignment = Alignment(horizontal='right')
tc = ws.cell(row=out_row, column=14, value=f'=SUM(N2:N{out_row-2})')
tc.font = Font(bold=True, color='FFFFFF')
tc.number_format = '"$"#,##0.00'
tc.fill = PatternFill('solid', fgColor='305496')

wb.save(XLSX)

print()
print('Consolidation summary:')
print(f'{"Section":<10} {"Original":<12} {"Consolidated":<14}')
print('-' * 40)
section_orig = {s: 0 for s, _, _ in sections}
section_new = {s: 0 for s, _, _ in consolidated}
for s, _, d in sections: section_orig[s] = len(d)
for s, _, d in consolidated: section_new[s] = len(d)
for s in section_orig:
    print(f'{s:<10} {section_orig[s]:<12} {section_new[s]:<14}')
print(f'{"TOTAL":<10} {orig_count:<12} {new_count:<14}')

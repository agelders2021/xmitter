"""Generate Resistor_Sourcing.xlsx -- shopping reference for the xmitter build.

Five sheets:
  1. Strategy        -- overview of the three-tier sourcing approach
  2. Kit Coverage    -- 63 values in the typical 1280-pc Amazon kit and
                        which schematic positions each one covers
  3. DigiKey Metal Film  -- the 5 values the kit doesn't have that ARE
                            critical for impedance/gain setting
  4. DigiKey Specialty   -- carbon comp, wirewound, metal oxide that no
                            general assortment covers
  5. Resistor Inventory  -- every distinct value used in the xmitter
                            design with its purpose and how to source it
"""

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).resolve().parent.parent / "Documentation" / "Resistor_Sourcing.xlsx"

# ── Colors / styles ──────────────────────────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
GOOD_FILL      = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL      = PatternFill("solid", fgColor="FFF2CC")
CRITICAL_FILL  = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Sheet 1: Strategy ────────────────────────────────────────────────────
def write_strategy(ws):
    ws.title = "Strategy"
    ws.cell(1, 1, "Resistor Sourcing Strategy").font = Font(
        bold=True, size=18, color="1F3864")

    ws.cell(3, 1, "Three-tier approach to covering all ~150 resistor positions:").font = Font(
        italic=True, color="555555")

    tier_data = [
        ("Tier", "Source", "Approx Cost", "Coverage"),
        ("1", "1280-pc 1 % metal-film kit from Amazon (Aniann / Bojack / MCIGICM / Bojack -- all same content)",
         "$15-25",
         "~75 % of build (E12-ish, 10R to 1M, 1/4 W, 1 %)"),
        ("2", "Yageo MFR-25 metal-film 1 % singles from DigiKey",
         "$3-5",
         "5 specific values for impedance / gain setting that the kit misses"),
        ("3", "Specialty parts from DigiKey or Mouser: Ohmite carbon comp + Ohmite Brown Devil wirewound + metal oxide",
         "$25-35",
         "Plate stoppers, screen droppers, driver high-power positions"),
    ]
    base_row = 5
    for i, row in enumerate(tier_data):
        r = base_row + i
        for j, val in enumerate(row, start=1):
            c = ws.cell(r, j, val)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)
            if i == 0:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(11, 1, "Order Checklist").font = Font(bold=True, size=14,
                                                   color="1F3864")
    checklist = [
        "[ ]  Amazon: 1280-pc 1 % metal-film resistor kit ($15-25)",
        "[ ]  DigiKey: 5 specific metal-film values (see 'DigiKey Metal Film' sheet)",
        "[ ]  DigiKey: Ohmite OD series carbon comp 47 ohm 5 W (PA plate stoppers)",
        "[ ]  DigiKey: Ohmite Brown Devil 1 kohm 5 W wirewound (PA screen droppers)",
        "[ ]  DigiKey: Metal oxide 1 W (driver stage)",
        "[ ]  Optional but useful: 0.1 % precision pair for V_REF divider if you want exact 1.5 V threshold",
        "",
        "Total estimated cost: $45-65 covering every resistor in the build with spares.",
    ]
    for i, line in enumerate(checklist):
        c = ws.cell(13 + i, 1, line)
        c.font = Font(name="Consolas", size=10)
        c.alignment = Alignment(horizontal="left", vertical="top")

    # Column widths
    for col, w in enumerate([6, 50, 16, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row heights for the tier table
    for r in range(base_row + 1, base_row + len(tier_data)):
        ws.row_dimensions[r].height = 45


# ── Sheet 2: Kit Coverage ────────────────────────────────────────────────
KIT_VALUES = [
    "1", "2.2", "4.7", "5.6", "7.5", "8.2",
    "10", "15", "22", "27", "33", "39",
    "47", "56", "68", "75", "82",
    "100", "120", "150", "180", "220", "270",
    "330", "390", "470", "510", "680", "820",
    "1k", "1.5k", "2.2k", "3k", "3.9k", "4.7k", "5.6k", "6.8k", "7.5k", "8.2k",
    "10k", "15k", "22k", "33k", "39k", "47k", "56k", "68k", "75k", "82k",
    "100k", "150k", "180k", "220k", "330k", "470k", "560k", "680k",
    "1M", "1.5M", "2M", "3.3M", "4.7M", "5.6M", "10M",
]

# Map each kit value to whether it's used in the xmitter and where.
# "OK"     = used as-is from the kit
# "skip"   = skip / not used in this build
# "carbon" = also use a carbon-comp version instead of metal film here
KIT_USAGE = {
    "10":  ("OK",     "Cathode sense R_C (NB also need 1 W carbon comp version)"),
    "47":  ("carbon", "PA plate stoppers -- METAL FILM NOT SUITABLE here; use Ohmite OD carbon comp 5 W"),
    "100": ("OK",     "Driver grid stoppers (driver also needs 1 W version)"),
    "120": ("skip",   ""),
    "150": ("OK",     "Possible alt R_F"),
    "220": ("OK",     "PA grid stoppers (R_2, R_6); also 22k positions"),
    "270": ("OK",     "J310 source bias RS1"),
    "470": ("OK",     "Misc"),
    "1k":  ("OK",     "R_pad in keyer + grid_bias, driver screen dropper (also need 1 W wirewound version)"),
    "1.5k":("OK",     "LM4040 bias resistor; envelope LPF R13"),
    "2.2k":("OK",     "Close-enough alt for 2.32k V_REF divider"),
    "3k":  ("OK",     "V_REF divider top leg"),
    "3.9k":("OK",     "MC1496 load RLa, RLb"),
    "4.7k":("OK",     "LM393 R_PULL"),
    "6.8k":("OK",     "MC1496 R_BIAS"),
    "8.2k":("OK",     "Null PNP collector load (R5, R7, R9, R10)"),
    "10k": ("OK",     "R_G, R_GATE, R_S (in cathode monitor), many places"),
    "22k": ("OK",     "PA grid leak, R_GL, MC1496 pin 8 divider"),
    "47k": ("OK",     "LM7171 R_F (= 47k for gain 5.8), driver grid leak"),
    "100k":("OK",     "Op-amp + input bias"),
    "150k":("OK",     "Possible alt"),
    "220k":("OK",     "Misc"),
    "470k":("OK",     "LM393 R_HYST hysteresis"),
    "1M":  ("OK",     "Gate leak (J310), sim test loads"),
}
KIT_USAGE_DEFAULT = ("skip", "Not used in this build")


def write_kit_coverage(ws):
    ws.title = "Kit Coverage"
    ws.cell(1, 1, "1280-pc Amazon kit -- value-by-value coverage").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1, "Aniann / Bojack / MCIGICM / EEEEE: identical 63-value content.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Status", "Used for"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for i, val in enumerate(KIT_VALUES):
        r = 5 + i
        status, note = KIT_USAGE.get(val, KIT_USAGE_DEFAULT)
        cells = [val + " ohm" if "k" not in val and "M" not in val
                       else val.replace("k", " kohm").replace("M", " Mohm"),
                 status, note]
        for j, cv in enumerate(cells, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if j != 2 else "center",
                                    vertical="center", wrap_text=True)
        # color the row based on status
        fill = (GOOD_FILL if status == "OK"
                else WARN_FILL if status == "carbon"
                else None)
        if fill is not None:
            for j in range(1, 4):
                ws.cell(r, j).fill = fill

    # Column widths
    for col, w in enumerate([14, 12, 70], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Sheet 3: DigiKey Metal Film Missing ──────────────────────────────────
DIGIKEY_METAL_FILM = [
    # (value, digikey_search, qty, used_for, note)
    ("51 ohm",  "Yageo MFR-25FBF52-51R",  10,
     "MC1496 R1 R3 R4 -- signal port + carrier port impedance terminations",
     "Substituting 47 R or 56 R shifts MC1496 port impedance; 51 R is the design value"),
    ("62 ohm",  "Yageo MFR-25FBF52-62R",  10,
     "VFO 20 dB Pi pad shunt resistors RP1 RP3",
     "Pad attenuation depends on exact value; 56 R or 68 R drifts pad ~1.5 dB"),
    ("200 ohm", "Yageo MFR-25FBF52-200R", 10,
     "MC1496 Re (between pins 2-3) -- sets modulator gain",
     "Modulator gain = R_L / Re. 180 R or 220 R from kit shifts gain ~10 %"),
    ("240 ohm", "Yageo MFR-25FBF52-240R", 10,
     "VFO 20 dB Pi pad series resistor RP2",
     "Pad attenuation depends on exact value"),
    ("170 kohm","Yageo MFR-25FBF52-170K", 10,
     "OPA454 R_F in grid_bias -- sets bias transfer function gain",
     "V_out = V_DAC * 18 - 85 requires R_F / R_G = 17. Kit's 150 k or 220 k shifts gain too far"),
]


def write_digikey_metal_film(ws):
    ws.title = "DigiKey Metal Film"
    ws.cell(1, 1, "DigiKey order -- metal-film 1 % singles the kit doesn't have").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "These 5 values are impedance- or gain-critical; must be exact. "
            "About $2.50-4.00 total for 10 each.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "DigiKey search term", "Qty", "Used for", "Why exact value matters"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(DIGIKEY_METAL_FILM):
        r = 5 + i
        val, dk, qty, used, note = row
        cells = [val, dk, qty, used, note]
        for j, cv in enumerate(cells, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.fill = CRITICAL_FILL
            c.alignment = Alignment(
                horizontal="center" if j == 3 else "left",
                vertical="top", wrap_text=True)
            if j == 2:  # DigiKey part num -- monospace
                c.font = Font(name="Consolas", size=10)

    # Column widths
    for col, w in enumerate([10, 28, 6, 50, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(DIGIKEY_METAL_FILM)):
        ws.row_dimensions[r].height = 50


# ── Sheet 4: Specialty parts ─────────────────────────────────────────────
SPECIALTY = [
    # (category, value, part / search, qty, used_for, notes)
    ("Carbon comp 5 W", "47 ohm",  "Ohmite OD470JE or OD470KE",       4,
     "PA plate stoppers R14 R15 -- one per tube + spares",
     "Non-magnetic / non-inductive structure beats metal film at 14 MHz. Also handles pulse loads."),
    ("Carbon comp 2 W", "220 ohm", "Ohmite OF221JE",                  4,
     "PA grid stoppers R2 R6 (optional upgrade from metal film)",
     "Metal film works here but carbon comp is the conservative choice for tube grid stoppers."),
    ("Wirewound 5 W",   "1 kohm",  "Ohmite Brown Devil B5J1K0E",      4,
     "PA screen droppers R3 R7 -- one per tube + spares",
     "Non-inductive style (Ohmite Series E). 5 W handles screen current droop."),
    ("Metal oxide 1 W", "100 ohm", "Yageo FMP100JR-52-100R",          8,
     "Driver grid stoppers (driver subcircuit)",
     "Non-inductive metal oxide. Carbon comp would also be fine here."),
    ("Metal oxide 1 W", "1 kohm",  "Yageo FMP100JR-52-1K",            4,
     "Driver screen droppers R2 R6 in driver subcircuit",
     "Drives ~25 mA at 150 V; 1 W gives margin."),
    ("Metal film 1 W",  "10 ohm",  "Yageo PR01000101000JR500",        4,
     "Cathode sense resistor R_C in PA cathode monitor",
     "1 W gives 4x margin at 150 mA cathode current (0.225 W dissipation)."),
    ("Carbon comp 5 W", "0.5 ohm", "(any 5 W resistor, sim convergence only)", 1,
     "Balun primary Rser -- sim convergence ONLY; do NOT install in hardware",
     "0.5 ohm in series with the balun primary stops a singular matrix in ngspice. The real hardware has no series R."),
]


def write_specialty(ws):
    ws.title = "DigiKey Specialty"
    ws.cell(1, 1, "DigiKey / Mouser -- specialty parts no general assortment covers").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1, "Carbon composition, wirewound, metal oxide. "
                  "Tube-specific positions where metal film isn't right.").font = Font(
        italic=True, color="555555")

    headers = ["Type", "Value", "Part / DigiKey search", "Qty", "Used for", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(SPECIALTY):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.alignment = Alignment(
                horizontal="center" if j == 4 else "left",
                vertical="top", wrap_text=True)
            if j == 3:
                c.font = Font(name="Consolas", size=10)

    # Column widths
    for col, w in enumerate([18, 12, 32, 6, 45, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(SPECIALTY)):
        ws.row_dimensions[r].height = 45


# ── Sheet 5: Resistor Inventory by Use ───────────────────────────────────
INVENTORY = [
    # (value, schematic, ref, sourcing, qty)
    ("10 ohm",   "PA_subcircuit",    "R1, R5",         "1 W metal film", "DigiKey: Yageo PR01000101000JR500 (×4)"),
    ("47 ohm",  "PA_subcircuit",    "R14, R15",       "5 W carbon comp", "DigiKey: Ohmite OD470 (×4)"),
    ("51 ohm",  "keyer",            "R1, R3, R4",     "1/4 W metal film 1 %", "DigiKey: Yageo MFR-25FBF52-51R (×10)"),
    ("62 ohm",  "vfo_subcircuit",   "RP1, RP3",       "1/4 W metal film 1 %", "DigiKey: Yageo MFR-25FBF52-62R (×10)"),
    ("100 ohm", "Driver_subcircuit","R1, R3, R8, R9", "1 W metal oxide", "DigiKey: Yageo FMP100JR-52-100R (×8)"),
    ("200 ohm", "keyer",            "Re",             "1/4 W metal film 1 %", "DigiKey: Yageo MFR-25FBF52-200R (×10)"),
    ("220 ohm", "PA_subcircuit",    "R2, R6",         "2 W carbon comp (or metal film)", "Kit covers; DigiKey OF221JE if going carbon comp"),
    ("240 ohm", "vfo_subcircuit",   "RP2",            "1/4 W metal film 1 %", "DigiKey: Yageo MFR-25FBF52-240R (×10)"),
    ("270 ohm", "vfo_subcircuit",   "RS1",            "1/4 W metal film 1 %", "Kit covers"),
    ("470 ohm", "various",          "misc",           "1/4 W metal film 1 %", "Kit covers"),
    ("1 kohm",  "Driver_subcircuit","R2, R6",         "1 W metal oxide",      "DigiKey: Yageo FMP100JR-52-1K (×4)"),
    ("1 kohm",  "PA_subcircuit",    "R3, R7",         "5 W wirewound non-inductive", "DigiKey: Ohmite Brown Devil B5J1K0E (×4)"),
    ("1 kohm",  "grid_bias, others","R_pad, misc",    "1/4 W metal film 1 %", "Kit covers"),
    ("1.5 kohm","grid_bias",        "LM4040 bias",    "1/4 W metal film 1 %", "Kit covers"),
    ("1.5 kohm","keyer",            "R13",            "1/4 W metal film 1 %", "Kit covers"),
    ("2.2 kohm","-- alt for 2.32 k","V_REF divider",  "1/4 W metal film 1 %", "Kit covers; gives ~1.47 V instead of 1.5 V (trim pot covers)"),
    ("3 kohm",  "(V_REF top leg)",  "",               "1/4 W metal film 1 %", "Kit covers"),
    ("3.9 kohm","keyer",            "RLa, RLb",       "1/4 W metal film 1 %", "Kit covers"),
    ("4.7 kohm","cathode_monitor",  "R_PULL",         "1/4 W metal film 1 %", "Kit covers"),
    ("6.8 kohm","keyer",            "R2 (R_BIAS)",    "1/4 W metal film 1 %", "Kit covers"),
    ("8.2 kohm","keyer",            "R5, R7, R9, R10","1/4 W metal film 1 %", "Kit covers"),
    ("10 kohm", "many",             "R_G, R_GATE, R_S","1/4 W metal film 1 %", "Kit covers"),
    ("22 kohm", "PA, grid_bias",    "R4, R8, R_GL",   "1/4 W metal film 1 %", "Kit covers"),
    ("47 kohm", "Op_Amp_Out",       "RF (LM7171)",    "1/4 W metal film 1 %", "Kit covers"),
    ("47 kohm", "Driver",           "R4, R7",         "1/4 W metal film 1 %", "Kit covers"),
    ("100 kohm","Op_Amp_Out",       "R1 (op-amp + bias)","1/4 W metal film 1 %", "Kit covers"),
    ("170 kohm","grid_bias",        "R_F (OPA454)",   "1/4 W metal film 1 %", "DigiKey: Yageo MFR-25FBF52-170K (×10)"),
    ("470 kohm","cathode_monitor",  "R_HYST",         "1/4 W metal film 1 %", "Kit covers"),
    ("1 Mohm",  "vfo, grid_bias",   "RGL, R_LOAD test","1/4 W metal film 1 %", "Kit covers"),
]


def write_inventory(ws):
    ws.title = "Resistor Inventory"
    ws.cell(1, 1, "Resistor inventory by value, with source").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1, "Cross-reference of every distinct value, where used, what type, and where to buy.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Schematic", "Ref desigs", "Type required", "Source / part"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(INVENTORY):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)
            if j == 5 and ("DigiKey" in (cv or "")):
                c.fill = CRITICAL_FILL
                c.font = Font(name="Consolas", size=10)
            elif j == 5 and "Kit covers" in (cv or ""):
                c.fill = GOOD_FILL

    # Column widths
    for col, w in enumerate([12, 20, 22, 24, 55], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


# ── Main ─────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_strategy(wb.create_sheet())
    write_kit_coverage(wb.create_sheet())
    write_digikey_metal_film(wb.create_sheet())
    write_specialty(wb.create_sheet())
    write_inventory(wb.create_sheet())

    OUT.parent.mkdir(exist_ok=True, parents=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()

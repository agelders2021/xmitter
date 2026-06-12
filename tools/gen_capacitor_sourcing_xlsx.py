"""Generate Capacitor_Sourcing.xlsx -- shopping reference for the xmitter build.

Six sheets:
  1. Strategy
  2. Silver Mica / C0G HV  -- LPF, tank, RF coupling (precision, HV)
  3. Ceramic Bypass        -- NP0 and X7R general decoupling
  4. Film                   -- envelope LPF and any audio-grade positions
  5. Aluminum Electrolytic -- supply bulk caps
  6. Trimmer Caps          -- adjustable, for fine-tuning LPF / tank / interelectrode
  7. Capacitor Inventory   -- every distinct value used, where, what type

Skipped per request: variable tuning caps (PA tank C12, C13; ATU caps).
Trimmer caps are INCLUDED.
"""

import os
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


OUT = Path(__file__).resolve().parent.parent / "Documentation" / "Capacitor_Sourcing.xlsx"

HEADER_FILL    = PatternFill("solid", fgColor="4472C4")
GOOD_FILL      = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL      = PatternFill("solid", fgColor="FFF2CC")
CRITICAL_FILL  = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ── Sheet 1: Strategy ────────────────────────────────────────────────────
def write_strategy(ws):
    ws.title = "Strategy"
    ws.cell(1, 1, "Capacitor Sourcing Strategy").font = Font(
        bold=True, size=18, color="1F3864")

    ws.cell(3, 1,
            "Capacitors are harder to assortment-shop than resistors because "
            "dielectric and voltage rating matter as much as value. "
            "Three-tier approach by dielectric type:").font = Font(
        italic=True, color="555555")

    tier_data = [
        ("Tier", "Source", "Coverage"),
        ("1", "DigiKey / Mouser: Cornell-Dubilier silver mica (~$2-4 each) for ALL the LPF, tank, and HV coupling positions",
         "Silver mica is the right answer for 5 % tolerance HV RF caps. ~$60-80 for the dozen positions in the build."),
        ("2", "Amazon / DigiKey: a general 'ceramic capacitor assortment kit' for NP0 and X7R decoupling (50-100 V)",
         "100 nF / 10 nF / 1 nF decoupling caps everywhere. Buy a 100-piece variety pack for ~$10 or order singles."),
        ("3", "DigiKey: specialty -- WIMA MKT/MKP film for 680 nF envelope LPF; Nichicon/Panasonic electrolytics for bulk supply bypass; Sprague-Goodman or Murata trimmers for adjustable positions",
         "Five-ish unique parts; order singles. ~$15-20."),
    ]
    base_row = 5
    for i, row in enumerate(tier_data):
        r = base_row + i
        for j, val in enumerate(row, start=1):
            c = ws.cell(r, j, val)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)
            if i == 0:
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = HEADER_FILL
                c.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(11, 1, "Order Checklist").font = Font(
        bold=True, size=14, color="1F3864")
    checklist = [
        "[ ]  DigiKey: silver mica caps for LPF / tank / RF coupling (see 'Silver Mica' sheet)",
        "[ ]  DigiKey or Amazon: ceramic NP0 / X7R bypass kit OR individual singles (see 'Ceramic Bypass' sheet)",
        "[ ]  DigiKey: WIMA MKT 680 nF film cap for envelope LPF C3",
        "[ ]  DigiKey: Aluminum electrolytic bulk caps for supply bypass",
        "[ ]  DigiKey or Mouser: trimmer caps -- Sprague-Goodman or Murata (see 'Trimmers' sheet)",
        "",
        "Variables NOT included (user already has): PA tank C12/C13 variable; ATU variables.",
        "",
        "Total estimated cost: $90-130 covering every fixed cap and trimmer in the build with spares.",
    ]
    for i, line in enumerate(checklist):
        c = ws.cell(13 + i, 1, line)
        c.font = Font(name="Consolas", size=10)
        c.alignment = Alignment(horizontal="left", vertical="top")

    for col, w in enumerate([6, 60, 60], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(base_row + 1, base_row + len(tier_data)):
        ws.row_dimensions[r].height = 55


# ── Sheet 2: Silver Mica / C0G HV ────────────────────────────────────────
SILVER_MICA = [
    # (value, voltage, qty, used_for, digikey_search, notes)
    ("6 pF",   "1 kV",   4,
     "Driver output transformer C1, C2 (interwinding stray capacitance modelling; small fixed caps in series with primary)",
     "Cornell-Dubilier CD15FD6R0DO3F",
     "Tiny value; alternatively model with a trimmer (see Trimmers sheet)."),
    ("47 pF",  "500 V",  6,
     "Q2 tank C1a (if used); driver tank fine tune; spare",
     "CD15ED470JO3F",
     "Often used as a trim around the variable tank cap."),
    ("100 pF", "500 V",  6,
     "PA screen RF bypass C9, C11 (×2 + spares)",
     "CD15ED101JO3F",
     "Mounted at the screen pin of each 6146B socket."),
    ("220 pF", "500 V",  4,
     "VFO 7-pole LPF CF1, CF7 (×2 + spares). Stable across temp for filter precision.",
     "CD15ED221JO3F",
     "5 % tolerance is critical for filter passband. Could substitute C0G class 1 ceramic."),
    ("300 pF", "500 V",  4,
     "Output LPF C1, C3 (×2 + spares); also used in driver tank position",
     "CD15ED301JO3F",
     ""),
    ("330 pF", "100 V",  4,
     "Keyer C1 (carrier coupling to MC1496 pin 10)",
     "CD15ED331JO3F",
     "Low voltage OK here; ceramic C0G also fine since coupling is small-signal."),
    ("390 pF", "500 V",  4,
     "VFO 7-pole LPF CF3, CF5 (×2 + spares). Or substitute 360 + 27 pF parallel.",
     "CD15ED391JO3F",
     ""),
    ("470 pF", "500 V",  4,
     "Output LPF C2 (center shunt). Some specs called for 450 pF; 470 is standard.",
     "CD15ED471JO3F",
     "Or parallel 270 + 200 pF for 470 pF exactly."),
    ("1000 pF (1 nF)", "1 kV", 6,
     "Driver output transformer C3, C4 (grid coupling); various RF coupling positions",
     "CD15ED102JO3F",
     "1 kV ensures margin for the PA grid side of the transformer."),
    ("10 nF (0.01 uF)", "500 V", 12,
     "Driver C2, C4, C5 (cathode + screen bypass); PA C8, C10 (screen bypass); various RF bypass",
     "CD15ED103JO3F (or use lower-V if not on screen rail)",
     "Many positions; buy in bulk."),
]


def write_silver_mica(ws):
    ws.title = "Silver Mica HV"
    ws.cell(1, 1, "Silver Mica / C0G ceramic for LPF, tank, HV coupling").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "DigiKey: search 'Cornell-Dubilier silver mica' or use part numbers below.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Voltage", "Qty", "Used for", "DigiKey search / part", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(SILVER_MICA):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.fill = CRITICAL_FILL
            c.alignment = Alignment(
                horizontal="center" if j in (1, 2, 3) else "left",
                vertical="top", wrap_text=True)
            if j == 5:
                c.font = Font(name="Consolas", size=10)

    for col, w in enumerate([14, 10, 6, 45, 28, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(SILVER_MICA)):
        ws.row_dimensions[r].height = 55


# ── Sheet 3: Ceramic Bypass (NP0 / X7R) ──────────────────────────────────
CERAMIC_BYPASS = [
    # (value, dielectric, voltage, qty, used_for, notes)
    ("1 nF",  "X7R", "50 V",  8,
     "Cathode monitor C_FILT (anti-alias + RF reject)",
     "DigiKey: TDK / Murata 1206 or thru-hole; cheap"),
    ("10 nF (0.01 uF)", "NP0/C0G", "100 V", 10,
     "Cathode resistor RF bypass C_BYP at tube socket; misc 14 MHz bypass",
     "Cornell-Dubilier or KEMET; 1206 SMD or radial"),
    ("100 nF (0.1 uF)", "X7R", "50-100 V", 30,
     "LM7171 AC coupling C_IN, C_OUT (×4 per channel pair = 8); Op-amp decoupling at every IC supply pin; LM393 decoupling; PA screen pin supply local bypass; etc.",
     "Buy in bulk; Yageo CC1206ZRY5V or KEMET equivalents.  THIS IS THE WORKHORSE BYPASS."),
    ("100 nF (0.1 uF)", "X7R", "500 V",  4,
     "Keyer C2 (MC1496 pin 8 AC bypass)",
     "Higher-V rated for the +12 V keyer rail; ceramic radial."),
    ("1 uF",  "X7R or X5R", "25-50 V", 6,
     "Supply rail decoupling near LM7171, LM393, OPA1641 (parallel with 100 nF and 10 uF)",
     "Murata GRM series or similar"),
]


def write_ceramic(ws):
    ws.title = "Ceramic Bypass"
    ws.cell(1, 1, "Ceramic NP0 / X7R decoupling and bypass caps").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "Many positions; buy in bulk. NP0/C0G for RF, X7R for general decoupling, "
            "X5R OK for low-voltage power supply bypass.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Dielectric", "Voltage", "Qty", "Used for", "Source / notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(CERAMIC_BYPASS):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.fill = GOOD_FILL
            c.alignment = Alignment(
                horizontal="center" if j in (2, 3, 4) else "left",
                vertical="top", wrap_text=True)

    for col, w in enumerate([16, 12, 12, 6, 50, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(CERAMIC_BYPASS)):
        ws.row_dimensions[r].height = 55


# ── Sheet 4: Film ────────────────────────────────────────────────────────
FILM = [
    ("680 nF",  "Film MKT or MKP", "50 V", 2,
     "Envelope reconstruction LPF C3 (keyer) -- single-pole RC with R13 = 1.5 kohm, fc = 4.75 kHz",
     "DigiKey: WIMA MKS2C046801K00JSSD or similar. Film is required (NOT ceramic) because 680 nF X7R loses 60 % of its capacitance at half rated voltage, ruining the LPF cutoff."),
    ("1 uF",   "Film",            "50-100 V", 2,
     "Possible alternate envelope LPF cap if you want lower fc",
     "Optional alternate; same WIMA MKS2 series"),
]


def write_film(ws):
    ws.title = "Film"
    ws.cell(1, 1, "Film capacitors -- envelope LPF and any audio-grade positions").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "Film for envelope LPF is essential: X7R ceramic at 680 nF loses too much C with applied V.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Type", "Voltage", "Qty", "Used for", "Source / notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(FILM):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.fill = CRITICAL_FILL
            c.alignment = Alignment(
                horizontal="center" if j in (2, 3, 4) else "left",
                vertical="top", wrap_text=True)

    for col, w in enumerate([10, 18, 12, 6, 50, 55], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(FILM)):
        ws.row_dimensions[r].height = 55


# ── Sheet 5: Aluminum Electrolytic ───────────────────────────────────────
ELECTROLYTIC = [
    ("10 uF",  "Aluminum electrolytic", "25-35 V", 12,
     "Supply rail bulk decoupling: +12 V, -8.3 V, +5 V (parallel with 100 nF ceramic at each IC area)",
     "Nichicon UVR series, or Panasonic FC. Radial thru-hole, low ESR."),
    ("10 uF",  "Aluminum electrolytic", "100 V",   4,
     "+3.3 V rail bulk cap near cathode-monitor clamp diodes (absorbs fault transient)",
     "Higher V rating gives margin for transient excursions."),
    ("47 uF",  "Aluminum electrolytic", "25 V",    4,
     "+5 V rail bulk cap (used + spares)",
     "Optional; alternative to 10 uF bulk for higher hold-up time."),
    ("100 uF", "Aluminum electrolytic", "25-35 V", 4,
     "+12 V rail bulk cap",
     "Optional; alternative to 10 uF bulk."),
]


def write_electrolytic(ws):
    ws.title = "Electrolytic"
    ws.cell(1, 1, "Aluminum electrolytic -- supply bulk caps").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "Used in pairs with 100 nF X7R at every IC for combined HF + LF decoupling.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Type", "Voltage", "Qty", "Used for", "Source / notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(ELECTROLYTIC):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.fill = GOOD_FILL
            c.alignment = Alignment(
                horizontal="center" if j in (2, 3, 4) else "left",
                vertical="top", wrap_text=True)

    for col, w in enumerate([10, 22, 12, 6, 50, 55], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(ELECTROLYTIC)):
        ws.row_dimensions[r].height = 45


# ── Sheet 6: Trimmer Caps ────────────────────────────────────────────────
TRIMMERS = [
    # (value range, voltage, qty, used_for, part, notes)
    ("1-10 pF",  "100 V",  2,
     "PA interelectrode strays C6, C7 (0.24 pF nominal -- adjust to compensate for tube + socket strays)",
     "Sprague-Goodman GMC10100 or Murata TZB4Z100",
     "Small ceramic / glass trimmers; series with the actual cap or used as the cap itself."),
    ("3-30 pF",  "250 V",  4,
     "Driver tank fine adjust; VFO LPF trimming (one per inductor if you want precise LPF response); spares",
     "Sprague-Goodman GKG30015 (or eBay PCB-mount air variable)",
     "More common value; useful for any small RF tank tuning."),
    ("5-50 pF",  "500 V",  2,
     "Output LPF center cap tune (if you want exact 17.5 MHz cutoff after winding tolerance)",
     "Sprague-Goodman GAG50015 or air variable",
     "Higher V rated for the output LPF position."),
]


def write_trimmers(ws):
    ws.title = "Trimmers"
    ws.cell(1, 1, "Trimmer capacitors -- adjustable for fine LPF / tank tuning").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "Variable tuning caps (PA tank, ATU) NOT included -- already obtained. "
            "Trimmers are the small adjustable caps for compensating wind / temp / stray variation.").font = Font(
        italic=True, color="555555")

    headers = ["Range", "Voltage", "Qty", "Used for", "Part / search", "Notes"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(TRIMMERS):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.fill = WARN_FILL
            c.alignment = Alignment(
                horizontal="center" if j in (2, 3) else "left",
                vertical="top", wrap_text=True)
            if j == 5:
                c.font = Font(name="Consolas", size=10)

    for col, w in enumerate([12, 10, 6, 45, 30, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    for r in range(5, 5 + len(TRIMMERS)):
        ws.row_dimensions[r].height = 55


# ── Sheet 7: Inventory by Use ────────────────────────────────────────────
INVENTORY = [
    # (value, schematic, ref, type, source)
    ("0.24 pF", "PA_subcircuit",        "C6, C7",       "Trimmer 1-10 pF (used as variable)", "Trimmers sheet"),
    ("6 pF",    "Driver_output_xfmr",   "C1, C2",       "Silver mica 1 kV (or trimmer)",      "Silver Mica sheet"),
    ("47 pF",   "Driver_tank (alt)",    "(if used)",    "Silver mica 500 V",                  "Silver Mica sheet"),
    ("100 pF",  "PA_subcircuit",        "C9, C11",      "Silver mica 500 V",                  "Silver Mica sheet"),
    ("220 pF",  "vfo_subcircuit",       "CF1, CF7",     "Silver mica or C0G 500 V, 5 %",      "Silver Mica sheet"),
    ("300 pF",  "LPF_50ohm",            "C1, C3",       "Silver mica 500 V",                  "Silver Mica sheet"),
    ("330 pF",  "keyer",                "C1",           "C0G ceramic 100 V (or silver mica)", "Silver Mica sheet"),
    ("390 pF",  "vfo_subcircuit",       "CF3, CF5",     "Silver mica or C0G 500 V, 5 %",      "Silver Mica sheet"),
    ("470 pF",  "LPF_50ohm",            "C2",           "Silver mica 500 V",                  "Silver Mica sheet"),
    ("1 nF",    "Driver_output_xfmr",   "C3, C4",       "Silver mica 1 kV",                   "Silver Mica sheet"),
    ("1 nF",    "cathode_monitor",      "C_FILT",       "X7R ceramic 50 V",                   "Ceramic Bypass sheet"),
    ("10 nF",   "Driver_subcircuit",    "C2, C4, C5",   "NP0 ceramic 500 V (cathode + screen bypass)", "Silver Mica or Ceramic Bypass sheet"),
    ("10 nF",   "PA_subcircuit",        "C8, C10",      "NP0 ceramic 500 V (screen bypass)",  "Silver Mica or Ceramic Bypass sheet"),
    ("10 nF",   "cathode_monitor",      "C_BYP",        "NP0 ceramic 100 V (RF bypass at socket)", "Ceramic Bypass sheet"),
    ("100 nF",  "keyer",                "C2",           "Ceramic 500 V (MC1496 pin 8 bypass)", "Ceramic Bypass sheet"),
    ("100 nF",  "Op_Amp_Out",           "C_IN, C_OUT",  "X7R ceramic 50 V (AC coupling)",     "Ceramic Bypass sheet"),
    ("100 nF",  "all ICs",              "decoupling",   "X7R ceramic 50-100 V (at every IC supply pin)", "Ceramic Bypass sheet"),
    ("680 nF",  "keyer",                "C3",           "Film MKT or MKP 50 V (envelope LPF)", "Film sheet"),
    ("10 uF",   "all supply rails",     "bulk bypass",  "Aluminum electrolytic 25-35 V",      "Electrolytic sheet"),
    ("10 uF",   "cathode_monitor",      "+3.3 V bulk",  "Aluminum electrolytic 100 V (fault transient)", "Electrolytic sheet"),
]


def write_inventory(ws):
    ws.title = "Capacitor Inventory"
    ws.cell(1, 1, "Capacitor inventory by value, with source pointer").font = Font(
        bold=True, size=14, color="1F3864")
    ws.cell(2, 1,
            "Cross-reference of every distinct value, where used, what type, and which sheet has the order info.").font = Font(
        italic=True, color="555555")

    headers = ["Value", "Schematic", "Ref desigs", "Type required", "Sourcing sheet"]
    for j, h in enumerate(headers, start=1):
        c = ws.cell(4, j, h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BORDER

    for i, row in enumerate(INVENTORY):
        r = 5 + i
        for j, cv in enumerate(row, start=1):
            c = ws.cell(r, j, cv)
            c.border = BORDER
            c.alignment = Alignment(horizontal="left", vertical="top",
                                    wrap_text=True)

    for col, w in enumerate([12, 22, 18, 40, 28], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_strategy(wb.create_sheet())
    write_silver_mica(wb.create_sheet())
    write_ceramic(wb.create_sheet())
    write_film(wb.create_sheet())
    write_electrolytic(wb.create_sheet())
    write_trimmers(wb.create_sheet())
    write_inventory(wb.create_sheet())

    OUT.parent.mkdir(exist_ok=True, parents=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")


if __name__ == "__main__":
    main()

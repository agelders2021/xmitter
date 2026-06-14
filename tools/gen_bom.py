"""Generate xmitter BOM as a formatted Excel spreadsheet."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'Documentation', 'BOM.xlsx')

# Row fill colors keyed by part type
FILL = {
    'Resistor':      'FFF2CC',
    'Capacitor':     'DDEEFF',
    'Inductor':      'E2EFDA',
    'Transformer':   'C6EFCE',
    'Semiconductor': 'FCE4D6',
    'Tube':          'EAD1FF',
    'IC':            'FFD9B3',
    'Control':       'FFE4F0',  # front-panel encoders, switches, etc.
}

HEADERS = [
    'Part Type', 'Reference', 'Value', 'Notes',
    'Sheet(s)', 'Qty', 'Datasheet', 'Supplier', 'Obtained',
]

COL_WIDTHS = [14, 16, 14, 60, 14, 6, 28, 14, 10]

# (Part Type, Reference, Value, Notes, Sheet(s), Qty, Datasheet, Supplier, Obtained)
BOM = [
    # ── VFO ──────────────────────────────────────────────────────────────────
    # Signal flow: Si5351A CLK0 (8 mA) -> 20 dB Pi pad -> 7-pole Chebyshev LPF
    # (fc=17.5 MHz) -> 50 Ω load termination -> J310 source follower -> MC1496
    # carrier port.  Per 2026-06-08 PA-validation session, the pad was raised
    # from 6 dB to 20 dB and the LPF from 5-pole to 7-pole to clean up the
    # modulator carrier in linear-multiplier mode.
    ('IC',           'U_VFO',       'Adafruit Si5351A breakout',
     'Programmable clock generator board (Adafruit #2045 / SparkFun BOB-13580). '
     '25 MHz crystal, I²C addr 0x60. CLK0 drives the pi-pad input. Drive strength = 8 mA. '
     'Treat its CLK0 output as ill-defined ~50–85 Ω (CMOS, not a clean R) — the pad swamps it.',
     'VFO', 1, '', 'Adafruit / SparkFun', ''),
    ('Resistor',     'RS',          '50Ω',
     'Metal film 1%; Si5351 source termination — sits across CLK0 to GND so the pad sees a defined ~50 Ω regardless of chip impedance',
     'VFO', 1, '', '', ''),
    ('Resistor',     'RP1, RP3',    '62Ω',
     'Metal film 1%; 20 dB Pi-attenuator shunt arms (2 identical). E96 alt: 61.9 Ω',
     'VFO', 2, '', '', ''),
    ('Resistor',     'RP2',         '240Ω',
     'Metal film 1%; 20 dB Pi-attenuator series arm. E96 alt: 249 Ω',
     'VFO', 1, '', '', ''),
    ('Capacitor',    'CF1, CF7',    '220pF',
     'NP0/C0G or silver mica, 50 V, 5%; 7-pole 0.1 dB Chebyshev LPF end-shunt caps (2 identical)',
     'VFO', 2, '', '', ''),
    ('Capacitor',    'CF3, CF5',    '390pF',
     'NP0/C0G or silver mica, 50 V, 5%; 7-pole Chebyshev LPF inner-shunt caps (or 360 + 27 pF parallel)',
     'VFO', 2, '', '', ''),
    ('Inductor',     'LF2, LF6',    '647nH',
     'T-50-6 (Mix 6, yellow); 13T #22 AWG enamelled; 7-pole Chebyshev LPF outer series arms',
     'VFO', 2, '', 'Amidon / Kits&Parts', ''),
    ('Inductor',     'LF4',         '715nH',
     'T-50-6 (Mix 6, yellow); 14T #22 AWG enamelled; 7-pole Chebyshev LPF centre series arm',
     'VFO', 1, '', 'Amidon / Kits&Parts', ''),
    ('Resistor',     'RT',          '50Ω',
     'Metal film 1%; 7-pole LPF load termination — IS a discrete resistor (the J310 gate taps across it through CC1)',
     'VFO', 1, '', '', ''),
    ('Resistor',     'RGL',         '1MΩ',
     'J310 gate bias (across RT — electrically invisible at RF, sets DC operating point)',
     'VFO', 1, '', '', ''),
    ('Resistor',     'RS1',         '270Ω',
     'J310 source bias (sets quiescent current, output node feeds MC1496 carrier port)',
     'VFO', 1, '', '', ''),
    ('Capacitor',    'CC1',         '10nF',
     'NP0/C0G 50 V; J310 gate AC coupling from LPF output through RT',
     'VFO', 1, '', '', ''),
    ('Semiconductor','T1',          'J310',
     'N-ch JFET; source follower; TO-92 through-hole. Alt: 2N5485',
     'VFO', 1, 'jfet-j308-j309-j310-interfet.pdf', '', ''),

    # Front-panel frequency tuning — drives the MCU which programmes the Si5351
    ('Control',      'ENC_FREQ',    'MBL-600-100P-5L',
     '100 PPR optical rotary encoder; +5V Vcc; AM26LS31 differential line-driver output (A, Ā, B, B̄). '
     'Cannot drive ESP32-S3 GPIOs directly — feeds U_ENC_RX (see below). MBLKJ industrial encoder.',
     'VFO', 1, 'rotary encoder datasheet.jpg', '', ''),
    ('IC',           'U_ENC_RX',    'AM26LS32ACN',
     'Quad differential line receiver (DIP-16), RS-422; pairs with the encoder\'s AM26LS31 outputs. '
     'Uses 2 of 4 receivers (A/Ā and B/B̄ pairs) and presents 3.3 V CMOS to the MCU. '
     'Alternatives: DS26C32AN (pin-compatible). Single-ended fallback if only A/B are used: '
     '74LVC2G17 dual buffer/level shifter (SOT-23-8) — loses differential noise rejection.',
     'VFO', 1, '', '', ''),

    # ── BUFFER / KEYER ───────────────────────────────────────────────────────
    # MC1496 used as a 4-quadrant multiplier (not switching mode): carrier port
    # is driven low (~32 mV pk after the upstream 20 dB pad), so harmonics at
    # the modulator output are dominated by carrier spectrum (clean) rather
    # than internal switching. The keyer DAC modulates pin 1 (signal port +);
    # AGC loop balances against pin 4. PNP nullers slam pin 1 / pin 4 high to
    # force the null condition during transmit fault / IDLE.
    ('IC',           'U1',          'MC1496 / LM1496',
     'Gilbert-cell balanced modulator used as a 4-quadrant analog multiplier; DIP-14; '
     'key-up null (zero control differential) eliminates RF leakage',
     'Buffer/Keyer', 1, 'MC1496_VCA_Buffer_Schematic.pdf', '', ''),
    ('IC',           'A1',          'TL071',
     'AGC error amp / integrator; DIP-8. TL072 dual is pin-compatible if substituting',
     'Buffer/Keyer', 1, 'tl071.pdf', '', ''),
    ('IC',           'DAC',         'MCP4921',
     '12-bit SPI DAC; envelope control to MC1496 pin 1 via 1.5 kΩ + 680 nF reconstruction LPF; '
     'driven by ESP32-S3 over dedicated SPI bus (NOT shared with monitoring I²C)',
     'Buffer/Keyer', 1, 'MC4921.pdf', '', ''),
    ('Transformer',  'T1',          'FT37-43',
     'Broadband push-pull interstage; 10+10T primary bifilar, 10+10T secondary bifilar; '
     '#28AWG enamelled; Fair-Rite Mix 43 ferrite',
     'Buffer/Keyer', 1, '', '', ''),
    ('Inductor',     'L_RFC',       '100µH',
     'Molded RFC; +12V to T1 primary CT; push-pull RF currents cancel at CT — only DC '
     'flows, so standard SRF is OK. Alt: 15T on FT37-43 (~94 µH)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Semiconductor','D1',          '1N5711',
     'Schottky; RF peak detector — samples one grid through C_DET_IN for AGC error signal',
     'Buffer/Keyer', 1, '', '', ''),
    ('Semiconductor','DZ1',         '1N4738A',
     '8.2V zener (1 W); generates −8.2 V rail for MC1496 V− (pin 7) and TL071 V− from the '
     'grid-bias rail (−60 to −80 V); ≈4 mA through DZ1 at nominal bias',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_DROP',      '15kΩ 1W',
     'Dropping resistor from grid-bias rail (−60 to −80V) to DZ1 zener; ≤300 mW at −80 V',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_GL',        '22kΩ',
     'T1 secondary CT grid-leak to GND (or return to a fixed −bias rail for class-C drive)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'RGS_A, RGS_B','100Ω',
     'Grid-stopper resistors — one in series with each 12HG7 driver-tube grid',
     'Buffer/Keyer', 2, '', '', ''),
    ('Resistor',     'R_SCALE1',    '100kΩ',
     'AGC divider upper leg: TL071 output → MC1496 signal port (pin 4)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_SCALE2',    '5.1kΩ',
     'AGC divider lower leg (4.85% ratio); keeps MC1496 control differential ≤ ±100 mV in the linear range',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_SIG_P',     '10kΩ',
     'MC1496 signal-port pin 1 reference to GND (sits at 0 V differential against the AGC node)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_INJ1, R_INJ2', '8.2kΩ',
     'Digital carrier-null injection (NEW in 2026-06-08 redesign): PNP collectors → MC1496 pins 1 and 4. '
     'Replaces the old 330 kΩ resistors — value chosen to balance the null PNP collector load to VEE',
     'Buffer/Keyer', 2, '', '', ''),
    ('Resistor',     'R_BIAS',      '6.8kΩ',
     'MC1496 bias pin 5 to GND; sets Iee ≈ 1 mA per AN531 (internal 500 Ω + this 6.8 kΩ)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_RE',        '200Ω',
     'MC1496 emitter degeneration between pins 2 and 3 (NOT to VEE). '
     'Reduced from 1 kΩ to 200 Ω in the 2026-06-08 redesign so modulator gain compensates for '
     'the lower carrier amplitude (the upstream 20 dB pad). Sets Av = RL/RE = 3.9 k / 200 ≈ 19.5',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_PIN8_TOP, R_PIN8_BOT', '22kΩ',
     'Carrier-port pin 8 bias divider, +12 V → pin 8 → GND. Raised from 1 kΩ to 22 kΩ in the '
     '2026-06-08 redesign so the standing current through the divider dropped from ~6 mA to ~0.3 mA '
     '(no functional change to pin 8 voltage)',
     'Buffer/Keyer', 2, '', '', ''),
    ('Resistor',     'R_INT',       '100kΩ',
     'TL071 inverting-input resistor (V_DET → A1 −)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_FB',        '470kΩ',
     'TL071 anti-windup / DC-gain feedback',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_DET',       '330kΩ',
     'Detector-diode trickle bias from +12 V; keeps the 1N5711 linear at low signal levels',
     'Buffer/Keyer', 1, '', '', ''),
    ('Resistor',     'R_ENV_LPF',   '1.5kΩ',
     'Envelope DAC reconstruction LPF series resistor (MCP4921 OUT → MC1496 pin 1). '
     'With R4 = 51 Ω to GND on pin 1 sets the DC divider (108 mV peak at full DAC code) and forms '
     'a ~4.75 kHz LPF with C_ENV_LPF. Bring-up note: see cw_envelope_keyer.md §"Hardware notes"',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_ENV_LPF',   '680nF',
     'Envelope LPF shunt cap (film MKT/MKP, 50 V). Forms the single-pole reconstruction filter '
     'with R_ENV_LPF and grounds pin 1 at 14 MHz (|Z| ≈ 0.017 Ω) — carrier isolation',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C1',          '330pF',
     'NP0/C0G 100 V; carrier coupling from Q1S follower output to MC1496 pin 10. '
     'Not used as an attenuator any more (upstream 20 dB pad does that); value unchanged from earlier rev',
     'Buffer/Keyer', 1, '', '', ''),
    ('Semiconductor','QT1, QT2',    '2N3906',
     'PNP small-signal switches (TO-92) driving the digital carrier-null injection into pins 1 and 4. '
     'Alt: BC557. One per side',
     'Buffer/Keyer', 2, '', '', ''),
    ('Resistor',     'R_PNP_C1, R_PNP_C2', '8.2kΩ',
     'PNP collector load to VEE (−8.2 V); pulls the injection node low when the PNP is off',
     'Buffer/Keyer', 2, '', '', ''),
    ('Resistor',     'R_PNP_E1, R_PNP_E2', '2.7kΩ',
     'PNP emitter to +5 V; sets PNP collector current during null injection',
     'Buffer/Keyer', 2, '', '', ''),
    ('Capacitor',    'C_INT',       '1µF film',
     'TL071 integrator cap; film type for low leakage / drift',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_DET',       '10nF film',
     'Detector smoothing cap (across the 1N5711 + R_DET node)',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_DET_IN',    '1000pF NP0',
     'Detector coupling from grid sample to 1N5711',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_GL',        '10nF NP0',
     'RF bypass across T1 secondary-CT grid-leak resistor R_GL',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_Z1',        '10µF electrolytic',
     '−8.2 V rail bulk bypass',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_Z2',        '100nF ceramic',
     '−8.2 V rail HF bypass',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_MC_CARR',   '100nF ceramic',
     'MC1496 carrier port − bypass (pin 8 to GND); keep lead short, close to the IC',
     'Buffer/Keyer', 1, '', '', ''),
    ('Capacitor',    'C_MC_VEE',    '100nF ceramic',
     'MC1496 VEE bypass (pin 7 to GND)',
     'Buffer/Keyer', 1, '', '', ''),

    # ── POST-KEYER AMP (LM7171) ──────────────────────────────────────────────
    # NEW in 2026-06-08 redesign.  The MC1496 now runs linear with the lower
    # 32 mV-pk carrier, so its differential output is only ~1.4 V p-p — well
    # below the 8 V p-p the 12HG7 driver wants.  One LM7171 per side at gain
    # ~5.8 brings the drive back to spec.  Powered from the existing +12 V /
    # −8.3 V rails (20.3 V total, inside the LM7171's ±18 V abs-max).
    ('IC',           'U2, U3',      'LM7171AIN',
     'High-speed voltage-feedback op-amp (200 MHz GBW); DIP-8 or SOIC-8. '
     'One per differential side. TL072 (3 MHz) is too slow — must use LM7171 or faster '
     '(AD8055, LMH6610, AD8009 all work as drop-ins for higher-gain variants)',
     'Post-Amp', 2, '', '', ''),
    ('Resistor',     'R_F_2, R_F_3', '47kΩ',
     'LM7171 feedback resistor; gain = 1 + R_F/R_G = 5.8. One per channel',
     'Post-Amp', 2, '', '', ''),
    ('Resistor',     'R_G_2, R_G_3', '10kΩ',
     'LM7171 inverting-input leg to GND; one per channel',
     'Post-Amp', 2, '', '', ''),
    ('Resistor',     'R_B_2, R_B_3', '100kΩ',
     'LM7171 + input bias to GND; one per channel',
     'Post-Amp', 2, '', '', ''),
    ('Capacitor',    'C_IN_2, C_IN_3', '100nF X7R',
     'AC-couple from MC1496 OUT (+10 V DC quiescent) to op-amp + input; one per channel',
     'Post-Amp', 2, '', '', ''),
    ('Capacitor',    'C_OUT_2, C_OUT_3', '100nF X7R',
     'AC-couple from op-amp output to 12HG7 cathode-input (grounded-grid stage); one per channel',
     'Post-Amp', 2, '', '', ''),
    ('Capacitor',    'C_BP_LM_BULK', '10µF aluminum',
     'LM7171 supply bulk bypass — one across +12 V → GND, one across −8.3 V → GND, ≤10 mm from supply pins',
     'Post-Amp', 2, '', '', ''),
    ('Capacitor',    'C_BP_LM_HF', '100nF ceramic',
     'LM7171 supply HF bypass — one per rail (+12 V and −8.3 V), ≤10 mm from supply pins',
     'Post-Amp', 2, '', '', ''),

    # ── DRIVER ───────────────────────────────────────────────────────────────
    ('Tube',         '2× 12HG7',    '12HG7',
     'Push-pull driver; 150V B+; push-pull pair',
     'Driver', 2, '12HG7.pdf', '', ''),
    ('Resistor',     'R2, R6',      '1kΩ',
     'Screen resistors',
     'Driver', 2, '', '', ''),
    ('Resistor',     'R4, R7',      '47kΩ',
     'Grid bias resistors',
     'Driver', 2, '', '', ''),
    ('Resistor',     'R1, R3, R8, R9', '100Ω',
     '4 identical resistors',
     'Driver', 4, '', '', ''),
    ('Capacitor',    'C2, C4, C5',  '10nF',
     '2× 10nF in parallel per position (shown as 0.01001µF in schematic); 3 positions = 6 caps total',
     'Driver', 6, '', '', ''),
    ('Capacitor',    'C6, C7',      '10nF',
     'Bypass caps',
     'Driver', 2, '', '', ''),
    ('Inductor',     'L1, L3',      '1µH',
     'Plate RFC; commercial molded choke; >50mA',
     'Driver', 2, '', '', ''),

    # ── PA ───────────────────────────────────────────────────────────────────
    ('Tube',         '2× 6146B',    '6146B',
     '600V B+, 200V screen, −70V grid; push-pull pair',
     'PA', 2, '6146b_big.pdf', '', ''),
    ('Resistor',     'R1, R5',      '10Ω',
     'Cathode resistors AND the cathode-current sense element (R_C in pa_cathode_monitor.md). '
     'Metal-film 1 W gives ~4× margin at 150 mA (0.225 W); 2 W carbon-comp also fine for headroom. '
     'Vishay PR01 or similar. Mount with C_BYP_CATH (0.01 µF NP0) directly at the tube-socket pin '
     '— the DC sense tap leaves the resistor top BEFORE that bypass cap',
     'PA', 2, '', '', ''),
    ('Capacitor',    'C_BYP_CATH1, C_BYP_CATH2', '0.01µF',
     'NP0 ceramic, 100 V, 1206 or smaller (low-ESL). Mounted DIRECTLY at the tube-socket cathode pin, '
     '≤5 mm leads. RF bypass that keeps 14 MHz out of the sense chain — one per tube',
     'PA', 2, '', '', ''),
    ('Resistor',     'R3, R7',      '1kΩ',
     'Screen resistors',
     'PA', 2, '', '', ''),
    ('Resistor',     'R4, R8',      '22kΩ',
     'Grid bias resistors',
     'PA', 2, '', '', ''),
    ('Resistor',     'R2, R6',      '220Ω',
     '',
     'PA', 2, '', '', ''),
    ('Resistor',     'R14, R15',    '47Ω',
     '',
     'PA', 2, '', '', ''),
    ('Capacitor',    'C6, C7',      '0.24pF',
     'Cak compensation; verify availability — very low value',
     'PA', 2, '', '', ''),
    ('Capacitor',    'C8, C10',     '10nF',
     'Bypass; must be rated for HV (≥1kV)',
     'PA', 2, '', '', ''),
    ('Capacitor',    'C9, C11',     '100pF',
     'Bypass; must be rated for HV (≥1kV)',
     'PA', 2, '', '', ''),
    ('Capacitor',    'C12, C13',    '33pF',
     'Plate tank; ganged variable cap 8–100pF; user has these',
     'PA', 2, '', '', ''),
    ('Inductor',     'L4, L5',      '1.71µH ea',
     'T-106-6 (Mix 6); center-tapped tank halves wound on ONE core with L6; ~19T each half; see L6 row',
     'PA', 1, '', 'Amidon / Kits&Parts', ''),
    ('Inductor',     'L6',          '0.27µH',
     'Tank link/output winding; same T-106-6 core as L4/L5; ~8T',
     'PA', 1, '', 'Amidon / Kits&Parts', ''),
    ('Inductor',     'L10, L11',    '1µH',
     'Screen/supply decoupling choke; commercial molded; HV-rated',
     'PA', 2, '', '', ''),

    # ── GRID BIAS (per-tube) ─────────────────────────────────────────────────
    # Each 6146B grid gets its own bias from MCP4728 DAC → OPA454 HV op-amp,
    # so firmware can fly the two-state IDLE / OPERATE setpoint (−90 V cutoff
    # for key-up, −50 V shallow class-C for key-down) and trim per-tube spread
    # in software.  Hardware-default supply is the −85 V rail so a firmware
    # hang parks the tubes safely in deep cutoff.
    # Schematic: xmitter_prj/grid_bias.sch  (instantiate per tube).
    ('IC',           'U_BIAS_DAC',  'MCP4728',
     'Quad 12-bit I²C DAC; one chip covers both tube grids (channels A & B) with C & D '
     'spare for future PA controls. Factory address 0x60 collides with Si5351 — order with '
     'an alternate address pre-programmed (e.g. 0x61), or sit on a second I²C peripheral. '
     'Safe-park: EEPROM startup code = 0 → −90 V (deep cutoff)',
     'Grid Bias', 1, 'MCP4725_2009.pdf', '', ''),
    ('IC',           'U_BIAS1, U_BIAS2', 'OPA454',
     'High-voltage op-amp (±100 V, ±50 mA); DIP-8 or SOIC-8. One per tube. '
     'Configured as inverter: DAC 0–5 V → −90 V (IDLE) … −50 V (OPERATE). '
     'Cheaper alt: discrete level-shifter with KSP44 PNP (400 V)',
     'Grid Bias', 2, 'opa454.pdf', '', ''),
    ('Resistor',     'R_BIAS_PAD1, R_BIAS_PAD2', '1kΩ',
     'Metal film 1%; DAC output → OPA454 + input series impedance (one per tube)',
     'Grid Bias', 2, '', '', ''),
    ('Resistor',     'R_BIAS_G1, R_BIAS_G2',     '10kΩ',
     'Metal film 1%; OPA454 inverting-input return to +5 V LM4040 reference (one per tube)',
     'Grid Bias', 2, '', '', ''),
    ('Resistor',     'R_BIAS_F1, R_BIAS_F2',     '170kΩ',
     'Metal film 1%; OPA454 feedback (sets gain ≈ 18 × DAC → −90…−50 V range; one per tube)',
     'Grid Bias', 2, '', '', ''),
    ('Resistor',     'R_BIAS_GL1, R_BIAS_GL2',   '22kΩ',
     'Metal film 1%; OPA454 output → 6146B grid (acts as both R_GL grid leak AND series '
     'isolation). One per tube — these REPLACE the role of the PA section R4 / R8',
     'Grid Bias', 2, '', '', ''),
    ('Resistor',     'R_BIAS_GATE1, R_BIAS_GATE2', '10kΩ',
     'Metal film 1%; Q_SLAM gate-drive resistor (CT_FAULT → Q_SLAM N-MOSFET gate; one per tube)',
     'Grid Bias', 2, '', '', ''),
    ('Semiconductor','Q_SLAM1, Q_SLAM2',         '2N7000',
     'N-channel MOSFET (TO-92); CT_FAULT bias-slam — pulls OPA454 + input to −85V rail '
     'on fault, slamming both tubes to deep cutoff in <1 µs. Alt: BSS138 (SOT-23). One per tube',
     'Grid Bias', 2, '', '', ''),
    ('IC',           'U_BIAS_REF',  'LM4040DIZ-5.0',
     'Precision 5.0 V shunt reference (TO-92); shared between both OPA454 stages. '
     'Bias resistor 1.5 kΩ 1/4 W from +12 V sets ~5 mA shunt current',
     'Grid Bias', 1, '', '', ''),
    ('Resistor',     'R_BIAS_REF',  '1.5kΩ',
     'Metal film 1%; LM4040 series resistor from +12 V (sets shunt current to ~5 mA)',
     'Grid Bias', 1, '', '', ''),
    ('Capacitor',    'C_BIAS_DEC_OPA', '100nF ceramic',
     'OPA454 supply bypass per rail (±100 V supply) — one per IC per rail, ≤10 mm from supply pins',
     'Grid Bias', 4, '', '', ''),
    ('Capacitor',    'C_BIAS_OUT1, C_BIAS_OUT2', '100nF',
     'NP0 200 V; OPA454 output to GND — RF bypass on the per-tube bias line. One per tube',
     'Grid Bias', 2, '', '', ''),
    # The −85 V rail itself: not a part per se — see HV supply chat. Captured here
    # so it doesn't get forgotten during BOM sourcing.
    ('IC',           'PSU_BIAS_NEG', '−85 V supply',
     'PA-grid bias rail, isolated DC-DC + TL431 shunt regulator (or small bench supply during bring-up). '
     'Defines hardware safe-park: with firmware off or hung, OPA454 + input pulls to this rail through '
     'R_BIAS_F → tubes parked in cutoff. Specification handled in the power-supply design chat',
     'Grid Bias', 1, '', '', ''),

    # ── CATHODE MONITOR (per-tube failsafe sense chain) ──────────────────────
    # 7-layer defense from a +600 V cathode-open fault to the MCU ADC pin, per
    # Documentation/pa_cathode_monitor.md.  Per-tube parts quantity = ×2.
    ('Inductor',     'F1_CATH1, F1_CATH2', 'MF-R010 PTC',
     'Bourns MF-R010 resettable PTC fuse (100 mA hold, 200 mA trip, V_max 60 V). '
     'Layer 1 — thermal slow-blow; opens within ~1 s on sustained fault; auto-resets. One per tube',
     'Cathode Monitor', 2, '', 'Bourns / Digi-Key', ''),
    ('Resistor',     'R_S_CATH1, R_S_CATH2', '10kΩ',
     'Metal film 1%, 1/4 W. Layer 2 — series current limiter; 600 V cathode-open fault → 60 mA into clamps. '
     'One per tube',
     'Cathode Monitor', 2, '', '', ''),
    ('Semiconductor','D_CATH1+, D_CATH1−, D_CATH2+, D_CATH2−', 'BAT54',
     'Schottky clamp diodes (SOT-23 or DO-35); D−=to GND, D+=to +3.3 V. Layer 3 — overvoltage clamp; '
     'survives the 60 mA fault transient until F1 trips. Alt: 1N5817 (higher current). Two per tube',
     'Cathode Monitor', 4, '', '', ''),
    ('Capacitor',    'C_FILT_CATH1, C_FILT_CATH2', '1nF',
     'X7R 50 V; Layer 3 anti-alias / RF rejection at the clamp node — forms 16 kHz LPF with R_S. '
     'One per tube',
     'Cathode Monitor', 2, '', '', ''),
    ('IC',           'U_CATH_BUF',  'OPA1642',
     'Dual FET-input op-amp (DIP-8) — Layer 4 ADC isolation, unity-gain buffer per tube. '
     'Has ±20 V differential input protection (extra safety even if D1/D2 fail). '
     'Alt: 2× OPA1641 single, or LMC6042 dual',
     'Cathode Monitor', 1, '', '', ''),
    ('IC',           'U_CATH_CMP',  'LM393',
     'Dual comparator (DIP-8) — Layer 5 fast hardware trip; one section per tube. '
     'Open-collector output → wired-OR latch → watchdog gate. Total trip <100 µs',
     'Cathode Monitor', 1, '', '', ''),
    ('IC',           'U_CATH_REF',  'LM4040DIZ-1.2',
     'Precision 1.225 V shunt reference (TO-92). Comparator trip threshold = 1.5 V '
     '(150 mA cathode); LM4040 + 1 kΩ trim pot sets the divider feeding LM393 (−) input. Shared',
     'Cathode Monitor', 1, '', '', ''),
    ('Resistor',     'R_HYST_CATH', '470kΩ',
     'Metal film 1%; LM393 positive-feedback hysteresis (50 mV trip-recovery gap)',
     'Cathode Monitor', 1, '', '', ''),
    ('Resistor',     'R_PULLUP_CATH', '4.7kΩ',
     'Metal film 1%; LM393 open-collector output pull-up to +3.3 V',
     'Cathode Monitor', 1, '', '', ''),
    ('Resistor',     'R_REF_TRIM',  '1kΩ trimpot',
     'Multi-turn trimmer (Bourns 3296 or similar); sets the LM4040 → LM393 (−) divider so trip lands '
     'at exactly 1.5 V_sense (150 mA cathode). Adjusted at bring-up calibration step 3',
     'Cathode Monitor', 1, '', '', ''),
    ('Capacitor',    'C_DEC_CATH',  '100nF ceramic',
     'X7R; supply decoupling — one at each IC supply pin (OPA1642, LM393, LM4040). Bulk supply',
     'Cathode Monitor', 5, '', '', ''),
    ('IC',           'U_CATH_ADC',  'ADS1115',
     '16-bit Δ-Σ I²C ADC, addr 0x48. Reads buffered V_sense from each tube at ~1 kHz (Layer 6 '
     'firmware soft trip + logging). Alt: built-in ESP32-S3 ADC if 12-bit (80 µA LSB) is acceptable',
     'Cathode Monitor', 1, '', '', ''),

    # ── WATCHDOG GATE (screen-voltage interrupter — Layer 5 actuator) ────────
    # When any fault path fires (LM393 hardware latch, esp_task_wdt GPIO, or
    # firmware fault::assert_fault()), Q_SCR_GATE opens and drops the +200 V
    # screen supply.  Tubes go dark in <100 µs.  Per pa_cathode_monitor.md
    # §"Watchdog gate".
    ('Semiconductor','Q_SCR_GATE',  'IRF9540',
     'P-channel power MOSFET (TO-220, V_DS=−100 V, I_D=−19 A). High-side switch in the +200 V '
     'screen supply. Default ON (R_GATE pulls to source); fault path pulls gate low via Q_GATE_DRV',
     'Watchdog Gate', 1, '', '', ''),
    ('Resistor',     'R_GATE_SCR',  '10kΩ',
     'Metal film 1/4 W; Q_SCR_GATE gate-to-source bias — keeps the gate high by default so the '
     'screen supply is normally enabled',
     'Watchdog Gate', 1, '', '', ''),
    ('Semiconductor','Q_GATE_DRV',  '2N3904',
     'NPN small-signal switch (TO-92); pulls Q_SCR_GATE gate low on fault. Driven by the wired-OR '
     'of: LM393 latch, MCU GPIO (FAULT_GATE_OUT), and esp_task_wdt interrupt output',
     'Watchdog Gate', 1, '', '', ''),
    ('Resistor',     'R_BASE_DRV',  '4.7kΩ',
     'Metal film 1/4 W; base-current limiter for Q_GATE_DRV',
     'Watchdog Gate', 1, '', '', ''),
    # Optional hardware SR latch — undecided pending pa_cathode_monitor.md open item.
    # Firmware-only latch via esp_task_wdt GPIO is simpler (~1 ms response) but slower
    # than a 74HC74 hardware latch (<100 µs response).  Adding here so it's not forgotten.
    ('IC',           'U_FAULT_LATCH', '74HC74 (optional)',
     'Dual D flip-flop (DIP-14); OPTIONAL hardware SR latch for the LM393 comparator outputs. '
     'Skip if firmware-only latch via esp_task_wdt is acceptable (~1 ms vs <100 µs response time). '
     'Decision deferred — see pa_cathode_monitor.md "Open items"',
     'Watchdog Gate', 1, '', '', ''),

    # ── DISPLAY (front-panel LCD) ────────────────────────────────────────────
    # Winstar WH2004A 20×4 character LCD, FSTN positive transflective, RGB
    # triple-color LED backlight.  HD44780-compatible controller.  Datasheets:
    # Documentation/Display/WH2004A-CFH-JT#.pdf and HD44780.pdf.
    #
    # Interface choice: I²C backpack + I²C RGB PWM driver — keeps the parallel
    # data bus off the limited ESP32-S3 GPIOs and avoids the 5V → 3.3V logic
    # mismatch (LCD V_IH = 0.7×VDD = 3.5 V, S3 GPIO ~3.0 V — out of spec direct).
    ('Control',      'LCD1',        'Winstar WH2004A-CFH-JT#',
     '20×4 character LCD, FSTN positive transflective, 6 o\'clock view, RGB triple-color LED backlight. '
     'VDD = 5 V (logic ~1.3 mA), backlight 42 mA R + 44 mA G + 47 mA B @ 5 V (~135 mA total full-white). '
     '18-pin connector: pins 1–15 are standard HD44780 + backlight anode; pins 16/17/18 = R/G/B cathodes',
     'Display', 1, 'Display/WH2004A-CFH-JT#.pdf', 'Winstar / Mouser / DigiKey', ''),
    ('IC',           'U_LCD_IF',    'PCF8574 I²C LCD backpack',
     'Generic "LCD I²C adapter" module — PCF8574 8-bit I²C expander + 16-pin LCD header + onboard '
     'contrast trimpot. Solders directly to pins 1–16 of LCD1. I²C addr 0x27 or 0x3F (solder-jumper '
     'selectable). Does NOT drive pins 16/17/18 (RGB cathodes) — those are handled by U_LCD_RGB',
     'Display', 1, '', 'Adafruit / SparkFun / generic', ''),
    ('IC',           'U_LCD_RGB',   'PCA9685 (Adafruit #815)',
     '16-channel I²C PWM driver, 12-bit per channel. Three channels drive the R/G/B cathodes on '
     'LCD pins 16/17/18 through R_LCD_BL_*. Default I²C addr 0x40 (no collision with Si5351 / MCP4728 '
     '/ ADS1115 / MAX17048 / PCF8574). 13 channels left free for future panel LEDs, indicators, etc. '
     'Alt: PCA9633 (4-ch, smaller) if PCB space is tight',
     'Display', 1, '', 'Adafruit / SparkFun', ''),
    ('Resistor',     'R_LCD_BL_R, R_LCD_BL_G, R_LCD_BL_B', '120Ω',
     'Metal film 1/4 W; current-limit resistor in series with each backlight cathode pin (16/17/18). '
     '120 Ω at V_LED ≈ 2 V gives ~25 mA per channel (below the 42–47 mA spec — extends lamp life; '
     'reduce to ~68 Ω for max brightness)',
     'Display', 3, '', '', ''),
    ('Capacitor',    'C_LCD_BYP',   '100nF ceramic',
     'X7R; VDD bypass at LCD VDD pin 2, ≤5 mm from pin',
     'Display', 1, '', '', ''),
    ('Capacitor',    'C_PCA_BYP',   '100nF ceramic',
     'X7R; VDD bypass at PCA9685, ≤5 mm from supply pin',
     'Display', 1, '', '', ''),
    # Connector — soldered to LCD bottom edge; choose pin header OR ribbon
    # depending on whether the LCD mounts on the same PCB or on a panel
    # cable.  18 pins because the WH2004A breaks out RGB separately from the
    # usual 16-pin "single LED" LCD.
    ('Control',      'J_LCD',       '18-pin 2.54mm header',
     'Single-row 0.1" pin header (18 contacts) for the LCD edge connector. Alt: 18-conductor ribbon '
     'cable + IDC headers if mounting LCD on a panel away from the main PCB',
     'Display', 1, '', '', ''),

    # ── MAINS INTERLOCK (boot/heartbeat fail-safe) ───────────────────────────
    # Hardware deadman switch: K_MAIN normally-open relay in series with the
    # AC mains feeding the HV / bias / screen / filament transformers.  Coil
    # is energised only while the MCU is actively pulsing GPIO MAINS_HEARTBEAT
    # at ≥5 Hz.  A 74HC4538 retriggerable monostable (R/C ≈ 200 ms) extends
    # each pulse — firmware hang / panic / brown-out → no more pulses → relay
    # drops → AC removed.  External 10 kΩ pull-down on the GPIO line guarantees
    # the relay is OFF before the MCU is even powered.  See pin_map.h
    # "Mains interlock" section.
    #
    # NOT in series with the low-voltage MCU supply — that comes up first
    # (USB or its own wallwart) so the firmware exists to start the heartbeat.
    ('IC',           'U_HBEAT',     '74HC4538',
     'Dual retriggerable monostable (DIP-16). One section used. R_RC + C_RC '
     'sets the retrigger window — pick R = 100 kΩ, C = 2.2 µF for ~220 ms hold. '
     'Output Q drives Q_HBEAT base. Pin 1 (A trigger) = MCU GPIO; pin 4 (Q) = relay drive',
     'Mains Interlock', 1, '', '', ''),
    ('Resistor',     'R_HBEAT_T',   '100kΩ',
     'Metal film 1%; 4538 timing resistor (pin 14 to VCC). With C_HBEAT_T = 2.2 µF gives ~220 ms hold',
     'Mains Interlock', 1, '', '', ''),
    ('Capacitor',    'C_HBEAT_T',   '2.2µF',
     'Film MKT 50 V; 4538 timing capacitor (pin 15 to GND). Tantalum / electrolytic OK if leakage low',
     'Mains Interlock', 1, '', '', ''),
    ('Resistor',     'R_HBEAT_PD',  '10kΩ',
     'Metal film 1/4 W; external pull-down on MAINS_HEARTBEAT GPIO trace. Guarantees relay OFF '
     'when MCU is unpowered, in reset, or removed',
     'Mains Interlock', 1, '', '', ''),
    ('Semiconductor','Q_HBEAT',     '2N3904',
     'NPN small-signal switch (TO-92); 4538 Q output → K_MAIN coil low-side. '
     'Alt: BSS138 SOT-23 if going SMT',
     'Mains Interlock', 1, '', '', ''),
    ('Resistor',     'R_HBEAT_B',   '4.7kΩ',
     'Metal film 1/4 W; Q_HBEAT base-current limiter from the 4538 output',
     'Mains Interlock', 1, '', '', ''),
    ('Semiconductor','D_HBEAT_FW',  '1N4148',
     'Flyback diode across K_MAIN coil — catches the kickback when Q_HBEAT switches off. '
     'Alt: 1N4001 for higher coil current relays',
     'Mains Interlock', 1, '', '', ''),
    ('IC',           'K_MAIN',      'Omron G7L-2A-T (or similar)',
     'AC mains power relay, SPST-NO (or DPST-NO for double-pole switching of L+N), '
     '12 V DC coil, contacts rated 250 V AC / 10 A min — sized for the rig\'s peak '
     'transformer inrush (factor ~10× steady-state). DIN-rail mount alternatives: '
     'Schrack RT, Omron MY series. Avoid sub-1 A "signal" relays — AC inrush will pit them',
     'Mains Interlock', 1, '', '', ''),
    ('Capacitor',    'C_K_BYP',     '100nF X7R',
     'K_MAIN coil supply bypass (12 V → GND), close to the relay coil',
     'Mains Interlock', 1, '', '', ''),
    # Notes — not parts, but record where this connects:
    # - K_MAIN contacts: in series with the HV mains feed BEFORE the plate / screen / bias /
    #   filament transformers.  NOT in series with the LV supply that powers the MCU.
    # - Coil supply (12 V DC): tapped from the LV supply rail or a small 12 V DIN-rail PSU
    #   dedicated to control circuits.
    # - The watchdog gate (Q_SCR_GATE in the Watchdog Gate section) and this mains interlock
    #   are independent layers: the gate is FAST (<100 µs) for cathode-current trips; the
    #   interlock is SLOW (~200 ms) for firmware failures.  Both belt-and-braces.

    # ── BALUN ────────────────────────────────────────────────────────────────
    ('Transformer',  'LP / LS',     '14µH / 2.33µH',
     'FT-114-61 ferrite (AL=75); 15T primary : 6T secondary; 300Ω balanced → 50Ω unbal; 6.25:1 Z ratio; good through 10m',
     'Balun', 1, '', 'Amidon / Kits&Parts', ''),

    # ── OUTPUT LPF ───────────────────────────────────────────────────────────
    ('Capacitor',    'C1, C3',      '300pF',
     'NP0/C0G 1–2%; 5-pole 0.5dB Chebyshev LPF fc=18MHz; end caps',
     'LPF (50Ω)', 2, '', '', ''),
    ('Capacitor',    'C2',          '450pF',
     'NP0/C0G 1–2%; 5-pole Chebyshev LPF centre cap',
     'LPF (50Ω)', 1, '', '', ''),
    ('Inductor',     'L1, L2',      '540nH',
     'T-68-6 (Mix 6, yellow); ~11T #26AWG; 5-pole Chebyshev LPF series arms',
     'LPF (50Ω)', 2, '', 'Amidon / Kits&Parts', ''),
]


def thin_border():
    s = Side(style='thin', color='BBBBBB')
    return Border(left=s, right=s, top=s, bottom=s)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOM'

    # Header row
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 20

    # Data rows
    data_align_wrap = Alignment(vertical='top', wrap_text=True)
    data_align_center = Alignment(horizontal='center', vertical='top')

    for row_idx, row in enumerate(BOM, start=2):
        part_type = row[0]
        color = FILL.get(part_type, 'FFFFFF')
        fill = PatternFill('solid', fgColor=color)

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            cell.border = thin_border()
            cell.font = Font(name='Calibri', size=10)
            if col_idx in (5, 6, 9):  # Sheet, Qty, Obtained — centre
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_wrap

        ws.row_dimensions[row_idx].height = 30

    # Auto-filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}1'

    # Freeze header
    ws.freeze_panes = 'A2'

    # Data validation for Obtained column (col 9)
    dv = DataValidation(
        type='list',
        formula1='"Yes,Ordered,No"',
        allow_blank=True,
        showDropDown=False,
    )
    ws.add_data_validation(dv)
    dv.sqref = f'I2:I{len(BOM) + 1}'

    wb.save(OUT_PATH)
    print(f'Saved: {os.path.abspath(OUT_PATH)}')


if __name__ == '__main__':
    build()

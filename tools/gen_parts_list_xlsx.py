"""Generate a per-sheet parts list from the QUCS-S schematics.

Walks xmitter_prj/*.sch, parses out R / C / L / SpLib / Sub / Vdc / Vac
components, and writes a multi-sheet Excel workbook with one tab per
schematic. Inductor / capacitor / resistor metadata (cores, voltages,
wattages, types) is filled in from the design docs where known.

Output: Documentation/Parts_List.xlsx
"""

import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


SCH_DIR = Path(__file__).resolve().parent.parent / "xmitter_prj"
OUT = Path(__file__).resolve().parent.parent / "Documentation" / "Parts_List.xlsx"

# Schematics to include, in logical signal-flow order. Skipping the
# scratch/test files and superseded ones.
SHEETS_IN_ORDER = [
    "xmitter.sch",
    "vfo_subcircuit.sch",
    "keyer.sch",
    "Op_Amp_Out.sch",
    "Driver_subcircuit.sch",
    "Driver_tank_subcircuit.sch",
    "Driver_output_transformer_subcircuit.sch",
    "PA_subcircuit.sch",
    "grid_bias.sch",
    "grid_bias_check.sch",
    "Balun_6to1_subcircuit.sch",
    "LPF_subcircuit.sch",
    "LPF_50ohm_subcircuit.sch",
    "ATU_subcircuit.sch",
    "Digital_subcircuit.sch",
]


# ── Per-sheet overrides for the "Specification" column ─────────────────────
# Keyed by (schematic_basename, reference_designator). Values populate the
# "Type / Spec" column with type, voltage rating, wattage, core+turns+wire,
# etc. taken from the design docs.

OVERRIDES = {
    # ── vfo_subcircuit.sch (Si5351 → pad → 7-pole LPF → J310 follower) ───
    ("vfo_subcircuit.sch", "RP1"): "Metal film 1/4 W 1 % 62 Ω (E24) or 61.9 Ω (E96)",
    ("vfo_subcircuit.sch", "RP2"): "Metal film 1/4 W 1 % 240 Ω (E24) or 249 Ω (E96)",
    ("vfo_subcircuit.sch", "RP3"): "Metal film 1/4 W 1 % 62 Ω (E24) or 61.9 Ω (E96)",
    ("vfo_subcircuit.sch", "RT"):  "Metal film 1/4 W 1 % 50 Ω (LPF load termination)",
    ("vfo_subcircuit.sch", "RGL"): "Metal film 1/4 W 1 % 1 MΩ (J310 gate leak)",
    ("vfo_subcircuit.sch", "RS"):  "Metal film 1/4 W 1 % 50 Ω (Si5351 source termination)",
    ("vfo_subcircuit.sch", "RS1"): "Metal film 1/4 W 1 % 270 Ω (J310 source bias)",
    ("vfo_subcircuit.sch", "CF1"): "Silver mica or C0G ceramic, 50 V, 5 %, 220 pF",
    ("vfo_subcircuit.sch", "CF3"): "Silver mica or C0G ceramic, 50 V, 5 %, 390 pF (or 360 + 27 pF parallel)",
    ("vfo_subcircuit.sch", "CF5"): "Silver mica or C0G ceramic, 50 V, 5 %, 390 pF",
    ("vfo_subcircuit.sch", "CF7"): "Silver mica or C0G ceramic, 50 V, 5 %, 220 pF",
    ("vfo_subcircuit.sch", "CC1"): "Ceramic NP0 50 V, 10 nF (J310 gate AC-coupling)",
    ("vfo_subcircuit.sch", "LF2"): "11 t #22 AWG enam. on T68-6 (yel., A_L=5.7 nH/N^2, OD 17.5 / ID 9.4 mm). Target 647 nH; 11 t gives ~690 nH -- spread turns slightly to trim down. Alt core: T50-6 @ 13 t.",
    ("vfo_subcircuit.sch", "LF4"): "11 t #22 AWG enam. on T68-6. Target 715 nH; 11 t gives ~690 nH -- bunch turns slightly to trim up. Alt core: T50-6 @ 14 t.",
    ("vfo_subcircuit.sch", "LF6"): "11 t #22 AWG enam. on T68-6. Same as LF2 (647 nH). Alt core: T50-6 @ 13 t.",
    # J310: through-hole TO-92, generic NJF
    ("vfo_subcircuit.sch", "JT1_J310"): "J310 NJF, TO-92. Alt: 2N5485.",

    # ── keyer.sch (MC1496 + null PNPs + carrier coupling + envelope LPF) ──
    ("keyer.sch", "RLa"):  "Metal film 1/4 W 1 % 3.9 kΩ (MC1496 load to +12 V)",
    ("keyer.sch", "RLb"):  "Metal film 1/4 W 1 % 3.9 kΩ (MC1496 load to +12 V)",
    ("keyer.sch", "R1"):   "Metal film 1/4 W 1 % 51 Ω (carrier port termination)",
    ("keyer.sch", "R2"):   "Metal film 1/4 W 1 % 6.8 kΩ (MC1496 R_BIAS, pin 5 to GND)",
    ("keyer.sch", "R3"):   "Metal film 1/4 W 1 % 51 Ω (signal port pin 4 bias to GND)",
    ("keyer.sch", "R4"):   "Metal film 1/4 W 1 % 51 Ω (signal port pin 1 bias to GND)",
    ("keyer.sch", "Re"):   "Metal film 1/4 W 1 % 200 Ω (MC1496 R_E, pins 2-3)",
    ("keyer.sch", "R5"):   "Metal film 1/4 W 1 % 8.2 kΩ (null PNP collector load to VEE)",
    ("keyer.sch", "R6"):   "Metal film 1/4 W 1 % 2.7 kΩ (null PNP emitter to +5 V)",
    ("keyer.sch", "R7"):   "Metal film 1/4 W 1 % 8.2 kΩ (null PNP collector load to VEE)",
    ("keyer.sch", "R8"):   "Metal film 1/4 W 1 % 2.7 kΩ (null PNP emitter to +5 V)",
    ("keyer.sch", "R9"):   "Metal film 1/4 W 1 % 8.2 kΩ (null injection R_INJ to MC1496 pin 1)",
    ("keyer.sch", "R10"):  "Metal film 1/4 W 1 % 8.2 kΩ (null injection R_INJ to MC1496 pin 4)",
    ("keyer.sch", "R11"):  "Metal film 1/4 W 1 % 22 kΩ (pin 8 bias divider, top)",
    ("keyer.sch", "R12"):  "Metal film 1/4 W 1 % 22 kΩ (pin 8 bias divider, bottom)",
    ("keyer.sch", "R13"):  "Metal film 1/4 W 1 % 1.5 kΩ (envelope LPF series; also signal-port pin 1 divider top)",
    ("keyer.sch", "C1"):   "Ceramic NP0 100 V, 330 pF, 5 % (carrier coupling to pin 10)",
    ("keyer.sch", "C2"):   "Ceramic NP0 100 V, 0.1 µF (pin 8 AC bypass)",
    ("keyer.sch", "C3"):   "Film MKT/MKP 50 V, 680 nF (envelope reconstruction LPF, to GND at pin 1)",
    ("keyer.sch", "QT1"):  "PNP small-signal: 2N3906 (TO-92) or BC557 (null injection driver)",
    ("keyer.sch", "QT2"):  "PNP small-signal: 2N3906 (TO-92) or BC557 (null injection driver)",
    ("keyer.sch", "X1"):   "MC1496 / LM1496 (DIP-14 or SO-14) -- balanced modulator",

    # ── Op_Amp_Out.sch (LM7171 unity-G ~6 voltage amp, two channels) ─────
    ("Op_Amp_Out.sch", "R1"):  "Metal film 1/4 W 1 % 100 kΩ (op-amp + input bias to GND)",
    ("Op_Amp_Out.sch", "RF"):  "Metal film 1/4 W 1 % 47 kΩ (LM7171 feedback; gain = 5.8)",
    ("Op_Amp_Out.sch", "RG"):  "Metal film 1/4 W 1 % 10 kΩ (LM7171 inverting input leg to GND)",
    ("Op_Amp_Out.sch", "C_IN"):  "Ceramic X7R 50 V, 100 nF (op-amp + input AC coupling)",
    ("Op_Amp_Out.sch", "C_OUT"): "Ceramic X7R 50 V, 100 nF (op-amp output AC coupling)",
    ("Op_Amp_Out.sch", "B_LM7171"): "LM7171 high-speed op-amp, DIP-8 or SOIC-8 (200 MHz GBW)",

    # ── Driver_subcircuit.sch (push-pull 12HG7 driver) ───────────────────
    ("Driver_subcircuit.sch", "R1"):  "Carbon comp or metal oxide 1 W 100 Ω (grid stopper)",
    ("Driver_subcircuit.sch", "R2"):  "Metal film 1 W 1 % 1 kΩ (screen dropper)",
    ("Driver_subcircuit.sch", "R3"):  "Carbon comp or metal oxide 1 W 100 Ω (grid stopper)",
    ("Driver_subcircuit.sch", "R4"):  "Metal film 1/4 W 1 % 47 kΩ (grid leak)",
    ("Driver_subcircuit.sch", "R6"):  "Metal film 1 W 1 % 1 kΩ (screen dropper)",
    ("Driver_subcircuit.sch", "R7"):  "Metal film 1/4 W 1 % 47 kΩ (grid leak)",
    ("Driver_subcircuit.sch", "R8"):  "Carbon comp or metal oxide 1 W 100 Ω (grid stopper)",
    ("Driver_subcircuit.sch", "R9"):  "Carbon comp or metal oxide 1 W 100 Ω (grid stopper)",
    ("Driver_subcircuit.sch", "C2"):  "Ceramic NP0 500 V, 10 nF (cathode bypass)",
    ("Driver_subcircuit.sch", "C4"):  "Ceramic NP0 500 V, 10 nF (cathode bypass)",
    ("Driver_subcircuit.sch", "C5"):  "Ceramic NP0 500 V, 10 nF (screen bypass)",
    ("Driver_subcircuit.sch", "C6"):  "Silver mica or C0G 500 V, 10 nF (grid coupling)",
    ("Driver_subcircuit.sch", "C7"):  "Silver mica or C0G 500 V, 10 nF (grid coupling)",
    ("Driver_subcircuit.sch", "L1"):  "1 µH RF choke -- molded or 12 t T37-2 (red), #26 AWG",
    ("Driver_subcircuit.sch", "L3"):  "1 µH RF choke -- molded or 12 t T37-2 (red), #26 AWG",
    ("Driver_subcircuit.sch", "V1"):  "Screen supply rail (+150 V regulated)",
    ("Driver_subcircuit.sch", "V5"):  "Screen supply rail (+150 V regulated)",
    ("Driver_subcircuit.sch", "V11"): "Cathode bias rail (+3 V; ideally from independent ref)",
    ("Driver_subcircuit.sch", "XSUB1"): "12HG7 video pentode, 9-pin Magnoval (B9D). Alt: 12HG7-A or 5763.",
    ("Driver_subcircuit.sch", "XSUB2"): "12HG7 video pentode, 9-pin Magnoval (B9D). Matched pair preferred.",

    # ── Driver_tank_subcircuit.sch ───────────────────────────────────────
    # (varies per design; if not in current sch, this block is harmless)

    # ── Driver_output_transformer_subcircuit.sch ─────────────────────────
    # All LQ1-4 windings on one core
    ("Driver_output_transformer_subcircuit.sch", "C1"): "Silver mica 1 kV, 6 pF (interwinding stray modelling)",
    ("Driver_output_transformer_subcircuit.sch", "C2"): "Silver mica 1 kV, 6 pF (interwinding stray modelling)",
    ("Driver_output_transformer_subcircuit.sch", "C3"): "Silver mica 1 kV, 1 nF (grid coupling cap)",
    ("Driver_output_transformer_subcircuit.sch", "C4"): "Silver mica 1 kV, 1 nF (grid coupling cap)",
    ("Driver_output_transformer_subcircuit.sch", "LQ1"): "Driver-side primary half-winding. Wind on a single common core (FT82-43 or T106-2).",
    ("Driver_output_transformer_subcircuit.sch", "LQ2"): "Driver-side primary half-winding (same core as LQ1).",
    ("Driver_output_transformer_subcircuit.sch", "LQ3"): "PA-grid secondary half-winding (same core).",
    ("Driver_output_transformer_subcircuit.sch", "LQ4"): "PA-grid secondary half-winding (same core).",

    # ── PA_subcircuit.sch (push-pull 6146B PA) ───────────────────────────
    ("PA_subcircuit.sch", "R1"):  "Carbon comp or metal-oxide 2 W 10 Ω (cathode resistor; sense across this)",
    ("PA_subcircuit.sch", "R5"):  "Carbon comp or metal-oxide 2 W 10 Ω (cathode resistor; sense across this)",
    ("PA_subcircuit.sch", "R3"):  "Wirewound 5 W 1 kΩ (screen dropper)",
    ("PA_subcircuit.sch", "R7"):  "Wirewound 5 W 1 kΩ (screen dropper)",
    ("PA_subcircuit.sch", "R4"):  "Metal film or carbon comp 1/4 W 22 kΩ (grid leak; only ~µA flow)",
    ("PA_subcircuit.sch", "R8"):  "Metal film or carbon comp 1/4 W 22 kΩ (grid leak)",
    ("PA_subcircuit.sch", "R14"): "Carbon comp 5 W 47 Ω (plate stopper; nonmagnetic)",
    ("PA_subcircuit.sch", "R15"): "Carbon comp 5 W 47 Ω (plate stopper; nonmagnetic)",
    ("PA_subcircuit.sch", "R2"):  "Carbon comp 2 W 220 Ω (grid stopper)",
    ("PA_subcircuit.sch", "R6"):  "Carbon comp 2 W 220 Ω (grid stopper)",
    ("PA_subcircuit.sch", "R10"): "Metal film 1/4 W 1 MΩ (sim load only -- do NOT populate in hardware)",
    ("PA_subcircuit.sch", "R11"): "Metal film 1/4 W 1 MΩ (sim load only -- do NOT populate in hardware)",
    ("PA_subcircuit.sch", "R17"): "External load model (300 Ω target via balun reflection); not a physical part",
    ("PA_subcircuit.sch", "C6"):  "Silver mica 2 kV 1 % 0.24 pF -- VERY small; consider trimmer or stray-only",
    ("PA_subcircuit.sch", "C7"):  "Silver mica 2 kV 1 % 0.24 pF -- VERY small; consider trimmer or stray-only",
    ("PA_subcircuit.sch", "C8"):  "Ceramic NP0 500 V, 10 nF (screen bypass)",
    ("PA_subcircuit.sch", "C9"):  "Silver mica 500 V, 100 pF (screen RF bypass)",
    ("PA_subcircuit.sch", "C10"): "Ceramic NP0 500 V, 10 nF (screen bypass)",
    ("PA_subcircuit.sch", "C11"): "Silver mica 500 V, 100 pF (screen RF bypass)",
    ("PA_subcircuit.sch", "C12"): "Variable tuning cap, 2 kV, ~10-50 pF (plate tank, ±33 pF nominal)",
    ("PA_subcircuit.sch", "C13"): "Variable tuning cap, 2 kV, ~10-50 pF (plate tank, ±33 pF nominal)",
    ("PA_subcircuit.sch", "L4"):  "Plate tank inductor half, 1.71 µH. Wind on T106-2 (red, 14 mm OD) or air-core ~20t #18 AWG enamelled, 20 mm diam.",
    ("PA_subcircuit.sch", "L5"):  "Plate tank inductor half, 1.71 µH. Same core as L4 (push-pull center-tap).",
    ("PA_subcircuit.sch", "L6"):  "Output link, 0.27 µH. ~4t #18 AWG, 20 mm diam. (or one turn on the L4/L5 chassis-mount tank).",
    ("PA_subcircuit.sch", "L10"): "1 µH plate decouple RFC -- molded, 1 W or 12 t T50-2 (red)",
    ("PA_subcircuit.sch", "L11"): "1 µH plate decouple RFC -- molded, 1 W or 12 t T50-2 (red)",
    ("PA_subcircuit.sch", "V1"):  "PA grid bias rail (currently fixed -70 V; will move to per-tube DAC + OPA454 via grid_bias subcircuit)",
    ("PA_subcircuit.sch", "V3"):  "PA grid bias rail (currently fixed -70 V; will move to per-tube DAC + OPA454 via grid_bias subcircuit)",
    ("PA_subcircuit.sch", "V2"):  "Screen supply rail (+200 V regulated; ~30 mA per tube)",
    ("PA_subcircuit.sch", "V4"):  "Screen supply rail (+200 V regulated; ~30 mA per tube)",
    ("PA_subcircuit.sch", "V5"):  "Plate supply rail (+600 V; ~150 mA per tube CW)",
    ("PA_subcircuit.sch", "XSUB1"): "6146B beam tetrode, 7-pin Octal (PA1). Alt: 6146A, 6883, 6293, 5894 (with re-bias).",
    ("PA_subcircuit.sch", "XX1"):  "6146B beam tetrode, 7-pin Octal (PA2). Matched pair preferred.",

    # ── grid_bias.sch (OPA454 per-tube bias control) ─────────────────────
    ("grid_bias.sch", "R_pad"):  "Metal film 1/4 W 1 % 1 kΩ (DAC -> (+) input impedance)",
    ("grid_bias.sch", "R_G"):    "Metal film 1/4 W 1 % 10 kΩ ((-) input return to +5 V LM4040)",
    ("grid_bias.sch", "R_F"):    "Metal film 1/4 W 1 % 170 kΩ (OPA454 feedback; sets gain 18)",
    ("grid_bias.sch", "R_GL"):   "Metal film 1/4 W 1 % 22 kΩ (grid leak / RFC to PA grid)",
    ("grid_bias.sch", "R_GATE"): "Metal film 1/4 W 1 % 10 kΩ (Q_SLAM gate driver)",
    ("grid_bias.sch", "LM40405"): "LM4040DIZ-5.0 (TO-92) + 1.5 kΩ 1/4 W bias resistor from +12 V",
    ("grid_bias.sch", "V2"):     "-85 V supply rail (small isolated DC-DC + TL431 shunt regulator)",
    ("grid_bias.sch", "V3"):     "+12 V system supply rail",
    ("grid_bias.sch", "X1"):     "OPA454 (DIP-8 or SOIC-8). 100 V op-supply max. ~$5.",
    ("grid_bias.sch", "Q_SLAM"): "2N7000 (TO-92) or BSS138 (SOT-23) N-MOSFET. Vth ~2 V.",

    # ── Balun_6to1_subcircuit.sch (PA output) ────────────────────────────
    ("Balun_6to1_subcircuit.sch", "R1"): "Carbon comp 5 W 0.5 Ω (sim convergence series R; 0 Ω in hardware)",
    ("Balun_6to1_subcircuit.sch", "LP"): "Balun primary 14 µH. ~17 turns on FT82-43 ferrite toroid bifilar with secondary.",
    ("Balun_6to1_subcircuit.sch", "LS"): "Balun secondary 2.33 µH. ~7 turns on the same FT82-43 core (6:1 impedance, ~2.45:1 turns).",

    # ── LPF_50ohm_subcircuit.sch (output LPF after balun) ────────────────
    ("LPF_50ohm_subcircuit.sch", "C1"): "Silver mica 500 V, 300 pF (output LPF shunt)",
    ("LPF_50ohm_subcircuit.sch", "C2"): "Silver mica 500 V, 450 pF (output LPF shunt center; or 470 pF)",
    ("LPF_50ohm_subcircuit.sch", "C3"): "Silver mica 500 V, 300 pF (output LPF shunt)",
    ("LPF_50ohm_subcircuit.sch", "L1"): "540 nH series inductor. ~13 t on T50-2 (red, μ=10) #20 AWG. Alt: T50-6.",
    ("LPF_50ohm_subcircuit.sch", "L2"): "540 nH series inductor. Same as L1.",
}


# ── Component-type heuristics for parts NOT in OVERRIDES ───────────────────

def default_spec(sch_name, comp_type, ref, value):
    """Best-guess type/voltage/wattage when no explicit override exists."""
    v = (value or "").strip()
    vlow = v.lower()
    if comp_type == "R":
        return f"Metal film 1/4 W 1 % {v} (default; check operating power)"
    if comp_type == "C":
        # Pick a sensible cap type based on value
        if "f" in vlow or "uf" in vlow or "μf" in vlow:
            return f"Ceramic X7R or electrolytic, voltage TBD, {v}"
        if "n" in vlow:
            return f"Ceramic NP0/X7R 50-100 V, {v}"
        return f"Ceramic NP0 or silver mica 50-500 V, {v}"
    if comp_type == "L":
        return f"Wind on toroid (core TBD) for {v}, or molded RFC"
    if comp_type == "L_SPICE":
        return f"Wind on toroid (core TBD) for {v}; see notes for core / turns / wire"
    if comp_type == "INDQ":
        return f"Q-modelled inductor ({v}); construction in transformer notes"
    if comp_type == "K_SPICE":
        return f"Coupling coefficient k = {v} (transformer winding pair, not a physical part)"
    if comp_type == "JFET":
        return f"JFET ({v})"
    if comp_type == "_BJT":
        return f"BJT ({v})"
    if comp_type == "_MOSFET":
        return f"MOSFET ({v})"
    if comp_type in ("Vdc", "Vac"):
        return f"Voltage source: {v} (rail or test source)"
    if comp_type == "SpLib":
        return f"SPICE library device: {v}"
    if comp_type == "Sub":
        return f"Sub-schematic instance: {v}"
    return v or ""


# ── Schematic parser ───────────────────────────────────────────────────────

# Match component lines like:
#   <R R1 1 200 430 -39 -55 0 0 "1 kOhm" 1 "26.85" 0 ...>
#   <C C1 1 ... "330pF" 1 ...>
#   <L L1 1 ... "1 uH" 1 ...>
#   <SpLib X1 1 ... "OPA454_5PIN" 0 ...>
#   <Sub SUB1 1 ... "grid_bias.sch" 0>
#   <Vdc V1 1 ... "12 V" 1>
#
# Capturing: component type, reference, first quoted string after the
# coordinate block (= the primary "value" / model name)
COMP_RE = re.compile(
    r'^\s*<(?P<type>R|C|L|L_SPICE|K_SPICE|INDQ|JFET|_BJT|_MOSFET|Vdc|Vac|Vrect|SpLib|Sub)\s+'
    r'(?P<ref>\S+)\s+\d+\s+'
    r'-?\d+\s+-?\d+\s+-?\d+\s+-?\d+\s+\d+\s+\d+\s+'
    r'"(?P<value>[^"]*)"',
)


def parse_sch(path: Path):
    """Yield (type, ref, value) for every component in a schematic."""
    text = path.read_text(errors="ignore")
    in_components = False
    for line in text.splitlines():
        line_stripped = line.strip()
        if line_stripped == "<Components>":
            in_components = True
            continue
        if line_stripped == "</Components>":
            in_components = False
            continue
        if not in_components:
            continue
        m = COMP_RE.match(line)
        if m:
            yield m.group("type"), m.group("ref"), m.group("value")


# ── Excel writer ───────────────────────────────────────────────────────────

HEADER = ["Reference", "Type", "Value", "Type / Spec (core, voltage, wattage)"]


def write_sheet(ws, sch_name, parts):
    """Populate one Excel sheet for one schematic."""
    ws.title = sch_name[:31]  # Excel sheet name limit

    # Title row
    ws.cell(row=1, column=1, value=sch_name).font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"{len(parts)} components")
    ws.cell(row=2, column=1).font = Font(italic=True, color="555555")

    # Header
    header_row = 4
    header_fill = PatternFill("solid", fgColor="e8eaf6")
    for col, label in enumerate(HEADER, start=1):
        c = ws.cell(row=header_row, column=col, value=label)
        c.font = Font(bold=True)
        c.fill = header_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows
    for i, (ctype, ref, value) in enumerate(parts, start=header_row + 1):
        spec = OVERRIDES.get((sch_name, ref))
        if spec is None:
            spec = default_spec(sch_name, ctype, ref, value)

        ws.cell(row=i, column=1, value=ref)
        ws.cell(row=i, column=2, value=ctype)
        ws.cell(row=i, column=3, value=value)
        ws.cell(row=i, column=4, value=spec)

        # Wrap long text in last column
        ws.cell(row=i, column=4).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )

    # Column widths
    widths = [14, 8, 18, 80]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w


def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default empty sheet

    total = 0
    for sch_name in SHEETS_IN_ORDER:
        path = SCH_DIR / sch_name
        if not path.exists():
            print(f"  (skip, missing: {sch_name})")
            continue
        parts = list(parse_sch(path))
        total += len(parts)
        ws = wb.create_sheet()
        write_sheet(ws, sch_name, parts)
        print(f"  {sch_name}: {len(parts)} parts")

    OUT.parent.mkdir(exist_ok=True, parents=True)
    wb.save(OUT)
    print(f"\nWrote {OUT}")
    print(f"Total parts across all sheets: {total}")


if __name__ == "__main__":
    main()

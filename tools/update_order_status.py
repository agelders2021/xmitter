"""Update VFO_Keyer_Order.xlsx: fill in On Hand, Ordered, and Ordered From
columns based on user's resistor stock list, current Mouser cart, and
eBay/Adafruit/surplus orders. Preserves all other user fills."""

import openpyxl
import shutil
import os

XLSX = 'Documentation/VFO_Keyer_Order.xlsx'
BAK = 'Documentation/VFO_Keyer_Order.xlsx.preupdate.bak'

if os.path.exists(BAK):
    os.remove(BAK)
shutil.copy(XLSX, BAK)
print(f'Backup: {BAK}')

# Resistor values on hand (per user 2026-06-18). All metal film 1/4W.
ON_HAND_RESISTORS_OHMS = {
    0, 1, 2.2, 4.7, 7.5, 10, 15, 22, 33, 39, 47, 56, 68, 100, 120, 150,
    220, 330, 390, 470, 510, 680,
    1000, 1500, 2000, 2200, 3000, 4700, 5100, 5600, 7500, 8200, 10000,
    15000, 22000, 33000, 47000, 56000, 68000, 75000, 100000,
    150000, 220000, 330000, 470000, 680000,
    1_000_000, 2_000_000, 4_700_000, 5_600_000,
}

# Map of Ref → (on_hand_qty, ordered_qty, source_label, note)
# Ref designators match the consolidated spreadsheet
ORDER_STATUS = {
    # VFO section
    'U1':           {'on_hand': 1, 'ordered': 1, 'from': 'Adafruit', 'note': 'Adafruit PID 5640'},
    'MTG':          {'ordered': 10, 'from': 'M', 'note': 'Essentra spacers (not threaded)'},
    'SCR':          {'ordered': 10, 'from': 'M'},
    'WIRE':         {'note': 'Use existing 24 AWG hookup wire stock'},
    'RS':           {'on_hand': 1, 'note': 'Have 0Ω in stock'},
    'C_coup':       {'ordered': 10, 'from': 'M', 'note': 'Vishay Y5P 10nF (fine for DC blocking)'},
    'RP1, RP3':     {'ordered': 10, 'from': 'M'},
    'RP2':          {'ordered': 10, 'from': 'M'},
    'CF1, CF7':     {'ordered_alt': 5, 'from_alt': 'Surplus',
                     'ordered': 4, 'from': 'M',
                     'note': 'Surplus 240pF × 5 (item 189) + Mouser 220pF × 4 — KEEP ONE, REMOVE OTHER'},
    'CF3, CF5':     {'ordered': 5, 'from': 'Surplus', 'note': 'Surplus item 192 (390pF × 5)'},
    'L2-6':         {'ordered': 14, 'from': 'eBay+Surplus',
                     'note': 'Amidon eBay (4) + surplus (10) = 14 T68-6 cores'},
    'WIRE2':        {'note': 'KitsAndParts #24 AWG 1/4 lb — separate order still needed'},
    'RT1':          {'ordered': 5, 'from': 'M', 'note': '50Ω termination'},
    'RGL':          {'on_hand': 1, 'note': 'Have 1MΩ in stock'},
    # J310 buffer
    'Q1':           {'ordered': 4, 'from': 'M'},
    'R6_V':         {'ordered': 5, 'from': 'M', 'note': '270Ω'},
    'CC1':          {'ordered_alt': 10, 'from_alt': 'M (shared with C_coup, Vishay Y5P 10nF)'},
    # Keyer section
    'U2':           {'ordered': 5, 'from': 'eBay', 'note': 'eBay 5pcs MC1496PG ($15.30)'},
    'XU2':          {'ordered': 4, 'from': 'M', 'note': 'Mill-Max premium machined'},
    'R1, R3, R4':   {'ordered': 6, 'from': 'M'},
    'R2':           {'ordered': 1, 'from': 'M', 'note': '6.8kΩ'},
    'RLa, RLb':     {'ordered': 6, 'from': 'M', 'note': 'Yageo 0.5% precision (better than spec)'},
    'R5, R7, R9, R10': {'on_hand': 4, 'ordered': 8, 'from': 'M',
                        'note': 'Have 8.2k in stock — can REMOVE from cart'},
    'R6_K, R8_K':   {'ordered': 6, 'from': 'M'},
    'R11, R12':     {'on_hand': 2, 'note': 'Have 22k in stock'},
    'R13':          {'on_hand': 1, 'note': 'Have 1.5k in stock'},
    'Re':           {'on_hand': 1, 'note': 'Have 2k in stock'},
    'C1':           {'ordered': 5, 'from': 'Surplus',
                     'note': 'Surplus item 191 (330pF silver mica × 5)'},
    'C2':           {'ordered_alt': 30, 'from_alt': 'M (shared with CB_IC, 0.1uF X7R × 30)'},
    'T1':           {'ordered': 5, 'from': 'M', 'note': '2N3906 PNP; match by Vbe'},
    'U3,4':         {'ordered': 4, 'from': 'M',
                     'note': 'LM7171BIN (B grade, not A — fine for AC-coupled RF use)'},
    'XU3,4':        {'ordered': 4, 'from': 'M', 'note': 'Mill-Max DIP-8 machined'},
    'C_IN, C_OUT':  {'ordered_alt': 30, 'from_alt': 'M (shared with CB_IC)'},
    'R_B':          {'on_hand': 2, 'note': 'Have 100k in stock — REMOVE from cart (qty 5)'},
    'R_F':          {'on_hand': 2, 'note': 'Have 47k in stock — REMOVE from cart (qty 5)'},
    'R_G':          {'on_hand': 2, 'ordered': 5, 'from': 'M',
                     'note': 'Have 10k 1% in stock; ordered 0.1% precision (optional upgrade)'},
    # Envelope DAC
    'U5':           {'ordered': 2, 'from': 'eBay',
                     'note': 'eBay 2pcs MCP4921 DIP-8 ($17.43) — REMOVE Mouser line 22'},
    'R_F_D':        {'on_hand': 1, 'note': 'Have 1.5k in stock (same as R13)'},
    'C_F':          {'ordered': 4, 'from': 'M', 'note': 'WIMA MKP4 680nF/100V'},
    # IC supply bypass
    'CB_IC':        {'ordered': 30, 'from': 'M', 'note': 'Kemet 0.1uF X7R × 30 (covers C2/C_IN/C_OUT/bypass)'},
    'CB_BLK':       {'ordered': 10, 'from': 'M',
                     'note': 'Aluminum electrolytic ECA-1EM100 (substituted from tantalum, saved $13)'},
    # Power
    'J_PWR':        {'ordered': 3, 'from': 'M',
                     'note': '3 × 2-pos (6 positions total; need 5 — OR swap to single 5-pos 1715061)'},
    'D1':           {'ordered': 8, 'from': 'M'},
    'R_DROP':       {'ordered': 5, 'from': 'M', 'note': '15kΩ 1W (different from 1/4W stock)'},
    # Connectors
    'J_RFOUT':      {'ordered': 2, 'from': 'M'},
    'J_SPI':        {'ordered': 2, 'from': 'M'},
    'J_PAD':        {'ordered': 2, 'from': 'M'},
    'WIRE3':        {'note': 'Use existing hookup wire stock'},
    # Test
    'PROBE':        {'note': 'Optional, defer'},
}

# Open workbook
wb = openpyxl.load_workbook(XLSX)
ws = wb['VFO + Keyer Order']

updated = 0
not_found = []
for row_idx in range(2, ws.max_row + 1):
    ref = ws.cell(row=row_idx, column=2).value
    if not ref:
        continue
    if ref in ORDER_STATUS:
        s = ORDER_STATUS[ref]
        # On Hand col 6
        if 'on_hand' in s:
            ws.cell(row=row_idx, column=6, value=s['on_hand'])
        # Ordered col 7 (only set if not already user-filled with bigger value)
        if 'ordered' in s:
            cur = ws.cell(row=row_idx, column=7).value
            if cur is None or cur == '' or (isinstance(cur, (int, float)) and cur < s['ordered']):
                ws.cell(row=row_idx, column=7, value=s['ordered'])
        # Ordered From col 9
        if 'from' in s:
            ws.cell(row=row_idx, column=9, value=s['from'])
        # Append note to Notes col 15
        if 'note' in s:
            existing = ws.cell(row=row_idx, column=15).value or ''
            new_note = s['note']
            if new_note not in str(existing):
                if existing.strip():
                    combined = str(existing).strip() + ' | ' + new_note
                else:
                    combined = new_note
                ws.cell(row=row_idx, column=15, value=combined)
        updated += 1
    else:
        not_found.append(ref)

wb.save(XLSX)
print(f'\nUpdated {updated} rows')
if not_found:
    print(f'\nRefs not in ORDER_STATUS map (may be section headers or new items):')
    for r in not_found:
        print(f'  {r}')

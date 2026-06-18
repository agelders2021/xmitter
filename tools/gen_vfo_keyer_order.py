"""Generate Documentation/VFO_Keyer_Order.xlsx — flat order list with check-off columns.

WARNING: this is a from-scratch bootstrap. Running it OVERWRITES any
On Hand / Ordered / Received / Ordered From / corrected P/N / price
entries the user has filled in. If you just want to add or change a
column structure without losing user data, write a migration script
that calls openpyxl.load_workbook() on the existing file and modifies
it in place (see tools/add_mfr_pn_column.py for the pattern).
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "VFO + Keyer Order"

headers = ['Section', 'Ref', 'Value / Description', 'Package', 'Qty',
           'On Hand', 'Ordered', 'Received', 'Ordered From',
           'Mfr P/N', 'Mouser P/N', 'Digi-Key P/N',
           'Unit $', 'Ext $', 'Notes']
col_widths = [13, 10, 38, 18, 5, 8, 8, 9, 13, 26, 26, 26, 8, 9, 40]

# Manufacturer P/Ns - populated only where unambiguous
MFR_PN = {
    'U1':    'Adafruit 5566',
    'U2':    'MC1496P',
    'Q1':    'J310',
    'T1':    '2N3906',
    'U3,4':  'LM7171AIN/NOPB',
    'U5':    'MCP4921-E/P',
    'D1':    '1N4738A',
    'CF1':   'CD15FD221JO3F',
    'CF3':   'CD15FD391JO3F',
    'CF5':   'CD15FD391JO3F',
    'CF7':   'CD15FD221JO3F',
    'C1':    'CD15FD331JO3F (silver mica option)',
    'CB_BLK':'Kemet T350F106K025AT',
    'C_F':   'Wima MKP4 680 nF / 63 V',
    'XU2':   'Mill-Max 110-43-314 (machined)',
    'XU3,4': 'Mill-Max 110-43-308 (machined)',
}

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill('solid', fgColor='305496')
section_fill = PatternFill('solid', fgColor='D9E1F2')
section_font = Font(bold=True, size=11)
check_fill = PatternFill('solid', fgColor='FFF2CC')
thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

for i, (h, w) in enumerate(zip(headers, col_widths), start=1):
    c = ws.cell(row=1, column=i, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 24
ws.freeze_panes = 'A2'

# (section, ref, value, package, qty, mouser_pn, digikey_pn, unit_price, notes)
parts = [
    # SI5351 + jumper hardware
    ('VFO', '', '--- Si5351 module + jumper hardware ---', '', None, '', '', '', ''),
    ('VFO', 'U1', 'Adafruit Si5351A STEMMA QT breakout', 'breakout', 1, '', '1528-5566-ND', 11.95, 'OWN this; Adafruit P/N 5566. Adafruit-only part.'),
    ('VFO', 'MTG', 'M2.5 x 10 mm nylon standoff (F-F)', 'mech', 4, '855-R30-1001003', '36-3475-ND', 0.60, 'Corner standoffs for Si5351'),
    ('VFO', 'SCR', 'M2.5 x 5 mm nylon pan-head screw', 'mech', 8, '534-29390', '36-29390-ND', 0.20, '2 per standoff'),
    ('VFO', 'WIRE', '24 AWG insulated hookup wire (asst colors)', 'wire', 1, '602-9314-4-100-NC025', '', 12.00, 'For 7 jumpers Si5351 to PCB + general use'),

    # VFO input chain
    ('VFO', '', '--- VFO input + pi pad + 7-pole LPF ---', '', None, '', '', '', ''),
    ('VFO', 'RS', '0 ohm 1/4W metal film (damping option)', 'axial THT', 1, '603-MFR-25FBF52-0R0', '', 0.10, 'Optional damper, default 0R'),
    ('VFO', 'C_coup', '10 nF NPO/C0G ceramic 100V 5%', 'radial disc D5mm', 1, '581-SR152A103JAR', '', 0.45, 'Si5351 CLK0 to pi pad coupling; blocks 1.65V offset'),
    ('VFO', 'RP1', '62 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-62R', '', 0.10, 'Pi pad input shunt'),
    ('VFO', 'RP2', '240 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-240R', '', 0.10, 'Pi pad series'),
    ('VFO', 'RP3', '62 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-62R', '', 0.10, 'Pi pad output shunt'),
    ('VFO', 'CF1', '220 pF silver mica 500V 5% (or 240 if surplus)', 'axial THT', 1, '598-CD15FD221JO3F', 'CD15FD221JO3F-ND', 1.85, 'LPF outer; surplus item 189 (240pF) acceptable'),
    ('VFO', 'CF3', '390 pF silver mica 500V 5%', 'axial THT', 1, '598-CD15FD391JO3F', 'CD15FD391JO3F-ND', 1.95, 'LPF; surplus item 192 exact match'),
    ('VFO', 'CF5', '390 pF silver mica 500V 5%', 'axial THT', 1, '598-CD15FD391JO3F', 'CD15FD391JO3F-ND', 1.95, 'LPF; surplus item 192 exact match'),
    ('VFO', 'CF7', '220 pF silver mica 500V 5% (or 240 if surplus)', 'axial THT', 1, '598-CD15FD221JO3F', 'CD15FD221JO3F-ND', 1.85, 'Same as CF1'),
    ('VFO', 'L2-6', 'T68-6 iron-powder toroid core', 'toroid', 5, '', '', 0.85, 'KitsAndParts.com only - neither Mouser nor DK stocks Micrometals cores'),
    ('VFO', 'WIRE2', '#24 AWG enameled magnet wire (1/4 lb spool)', 'spool', 1, '', '', 8.00, 'KitsAndParts.com - neither Mouser nor DK stocks ham magnet wire'),
    ('VFO', 'RT1', '50 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-49R9', '', 0.10, 'LPF output termination (49.9 ohm = nearest E96)'),
    ('VFO', 'RGL', '1 Mohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-1M', '', 0.10, 'MC1496 carrier-port grid-leak; ADD TO SCH'),

    # J310 buffer
    ('VFO', '', '--- J310 source follower buffer (optional, can jumper-bypass) ---', '', None, '', '', '', ''),
    ('VFO', 'Q1', 'J310 N-channel JFET', 'TO-92', 2, '512-J310', 'J310-ND', 0.85, '1 + 1 spare; static-sensitive'),
    ('VFO', 'R6_V', '270 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-270R', '', 0.10, 'J310 source resistor'),
    ('VFO', 'CC1', '10 nF NPO/C0G ceramic 100V', 'radial disc D5mm', 1, '581-SR152A103JAR', '', 0.45, 'J310 output coupling to MC1496'),

    # Keyer MC1496
    ('KEYER', '', '--- MC1496 keyer + bias network ---', '', None, '', '', '', ''),
    ('KEYER', 'U2', 'MC1496P balanced modulator', 'DIP-14', 2, '511-MC1496P', 'MC1496PG-ND', 2.45, '1 + 1 spare; static-sensitive'),
    ('KEYER', 'XU2', 'DIP-14 IC socket, machined pin', 'DIP-14 socket', 1, '575-1102873', 'AE9990-ND', 0.85, 'Mill-Max machined preferred over stamped'),
    ('KEYER', 'R1', '51 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-51R', '', 0.10, 'MC1496 carrier port termination'),
    ('KEYER', 'R2', '6.8 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-6K8', '', 0.10, 'MC1496 Iee setting (pin 5)'),
    ('KEYER', 'R3', '51 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-51R', '', 0.10, 'MC1496 modulator port'),
    ('KEYER', 'R4', '51 ohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-51R', '', 0.10, 'Modulator port + DAC injection shunt at pin 1'),
    ('KEYER', 'RLa', '3.9 kohm 1% metal film 1/4W (match w/ RLb)', 'axial THT', 1, '603-MFR-25FRF52-3K9', '', 0.10, 'Output load; order 4-5 and match by DMM'),
    ('KEYER', 'RLb', '3.9 kohm 1% metal film 1/4W (match w/ RLa)', 'axial THT', 1, '603-MFR-25FRF52-3K9', '', 0.10, 'See RLa note'),
    ('KEYER', 'R5', '8.2 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-8K2', '', 0.10, 'PNP T1 base (symmetry)'),
    ('KEYER', 'R7', '8.2 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-8K2', '', 0.10, 'PNP T2 base (symmetry; match R5)'),
    ('KEYER', 'R9', '8.2 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-8K2', '', 0.10, 'PNP null injection'),
    ('KEYER', 'R10', '8.2 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-8K2', '', 0.10, 'PNP null injection'),
    ('KEYER', 'R6_K', '2.7 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-2K7', '', 0.10, 'PNP T1 collector bias'),
    ('KEYER', 'R8_K', '2.7 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-2K7', '', 0.10, 'PNP T2 collector bias (match R6_K)'),
    ('KEYER', 'R11', '22 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-22K', '', 0.10, 'Carrier bias divider top'),
    ('KEYER', 'R12', '22 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-22K', '', 0.10, 'Carrier bias divider bottom'),
    ('KEYER', 'R13', '1.5 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-1K5', '', 0.10, 'Modulator input divider'),
    ('KEYER', 'Re', '2.0 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-2K', '', 0.10, 'MC1496 emitter degen (gain set)'),
    ('KEYER', 'C1', '330 pF NPO/C0G ceramic 100V (or silver mica)', 'radial disc D5mm', 1, '581-SR152A331JAR', '', 0.45, 'Carrier port coupling; surplus item 191 also fine'),
    ('KEYER', 'C2', '0.1 uF X7R ceramic 50V', 'radial 5mm', 2, '80-C320C104K5R', '399-9866-ND', 0.20, 'MC1496 pin 10 bypass + 1 spare'),
    ('KEYER', 'T1', '2N3906 PNP general-purpose', 'TO-92', 4, '610-2N3906', '2N3906FS-ND', 0.20, 'PNP null injection; buy 4 and pair by Vbe'),
    ('KEYER', 'T2', '2N3906 (covered above)', 'TO-92', 0, '', '', 0.00, 'Paired with T1 from the 4 ordered'),

    # LM7171 post-amps
    ('KEYER', '', '--- LM7171 differential post-amps (one per MC1496 output side) ---', '', None, '', '', '', ''),
    ('KEYER', 'U3,4', 'LM7171AIN high-speed op-amp', 'DIP-8', 3, '926-LM7171AIN/NOPB', 'LM7171AIN/NOPB-ND', 7.95, '2 needed + 1 spare'),
    ('KEYER', 'XU3,4', 'DIP-8 IC socket, machined pin', 'DIP-8 socket', 3, '575-1102871', 'AE9986-ND', 0.55, 'For LM7171 x 2 + MCP4921'),
    ('KEYER', 'C_IN', '100 nF X7R ceramic 50V', 'radial 5mm', 4, '80-C320C104K5R', '399-9866-ND', 0.20, '2 per LM7171 input AC couple'),
    ('KEYER', 'C_OUT', '100 nF X7R ceramic 50V', 'radial 5mm', 4, '80-C320C104K5R', '399-9866-ND', 0.20, '2 per LM7171 output AC couple'),
    ('KEYER', 'R_B', '100 kohm 1% metal film 1/4W', 'axial THT', 2, '603-MFR-25FRF52-100K', '', 0.10, 'LM7171 input bias'),
    ('KEYER', 'R_F', '47 kohm 1% metal film 1/4W', 'axial THT', 2, '603-MFR-25FRF52-47K', '', 0.10, 'LM7171 feedback (gain 5.8)'),
    ('KEYER', 'R_G', '10 kohm 1% metal film 1/4W', 'axial THT', 2, '603-MFR-25FRF52-10K', '', 0.10, 'LM7171 inverting leg to GND'),

    # Envelope DAC
    ('KEYER', '', '--- Envelope DAC + reconstruction LPF ---', '', None, '', '', '', ''),
    ('KEYER', 'U5', 'MCP4921-E/P 12-bit SPI DAC', 'DIP-8', 2, '579-MCP4921-E/P', 'MCP4921-E/P-ND', 3.25, '1 + 1 spare'),
    ('KEYER', 'R_F_D', '1.5 kohm 1% metal film 1/4W', 'axial THT', 1, '603-MFR-25FRF52-1K5', '', 0.10, 'DAC output to MC1496 pin 1'),
    ('KEYER', 'C_F', '680 nF polypropylene film (MKP) 63V+', 'box 5mm', 2, '871-B32921C3684K', '495-1395-ND', 1.20, 'DAC reconstruction LPF + 1 spare'),

    # IC supply bypass
    ('KEYER', '', '--- Per-IC supply bypass (one cap per supply pin) ---', '', None, '', '', '', ''),
    ('KEYER', 'CB_IC', '100 nF X7R ceramic 50V (per-IC bypass)', 'radial 5mm', 10, '80-C320C104K5R', '399-9866-ND', 0.20, '1 at MC1496 pin 7 + 4 at LM7171 supplies + 1 at MCP4921 + spares'),
    ('KEYER', 'CB_BLK', '10 uF tantalum 25V (per-rail bulk)', 'radial 5mm', 4, '80-T350F106K025AT', '399-3801-ND', 0.85, 'Bulk: +12V, -8.3V, +5V, +3.3V'),

    # Power input + bias derivation
    ('POWER', '', '--- Power input terminal block + -8.3V derivation from -90V ---', '', None, '', '', '', ''),
    ('POWER', 'J_PWR', '5-pos screw terminal block 5.08 mm pitch', 'THT 5pos', 1, '651-1715022', '277-1273-ND', 1.85, '+12V, +5V, -8.3V, -90V (bias), GND'),
    ('POWER', 'D1', '1N4738A 8.2V zener 1W', 'DO-41', 5, '512-1N4738A', '1N4738A-TPCT-ND', 0.18, '-8.3V derivation from -90V + spares'),
    ('POWER', 'R_DROP', '15 kohm 1W metal film resistor', 'axial 1W', 2, '603-PR01000101502JR500', '', 0.55, '-8.3V dropper from -90V rail + 1 spare'),

    # Connectors
    ('CONN', '', '--- Off-board connectors / headers ---', '', None, '', '', '', ''),
    ('CONN', 'J_RFOUT', '3-pin male header 0.1" pitch', 'THT 1x3', 1, '649-69190-403HLF', 'S1011E-03-ND', 0.50, 'RF output to driver: HOT+, HOT-, GND'),
    ('CONN', 'J_SPI', '5-pin male header 0.1" pitch', 'THT 1x5', 1, '649-69190-405HLF', 'S1011E-05-ND', 0.55, 'SPI to MCU: SCK, MOSI, CS, LDAC, GND'),
    ('CONN', 'J_PAD', '2-pin male header 0.1" pitch', 'THT 1x2', 1, '649-69190-402HLF', 'S1011E-02-ND', 0.45, 'Paddle dit/dah input (or use jumper wires)'),
    ('CONN', 'WIRE3', 'Hookup wire kit 22 AWG (6 colors)', 'wire', 1, '602-9314-4-100-WT005', '', 30.00, 'Inter-board + paddle wiring; OWN already?'),

    # Bench test
    ('TEST', '', '--- Bench bring-up extras (not on PCB) ---', '', None, '', '', '', ''),
    ('TEST', 'PROBE', '50 ohm BNC bulkhead jack', 'panel', 2, '523-31-221-RFX', 'ACX1230-ND', 6.50, 'Scope tap-offs at LPF out and MC1496 out (optional)'),
]

row = 2
section_color = {'VFO': 'D9E1F2', 'KEYER': 'E2EFDA', 'POWER': 'FFF2CC', 'CONN': 'FCE4D6', 'TEST': 'EDEDED'}
N_COLS = len(headers)  # 15
for section, ref, value, package, qty, mouser_pn, digikey_pn, unit, notes in parts:
    if ref == '' and qty is None:
        c = ws.cell(row=row, column=1, value=section)
        c.font = section_font
        c.fill = section_fill
        c2 = ws.cell(row=row, column=3, value=value)
        c2.font = Font(bold=True, italic=True)
        c2.fill = section_fill
        for col in range(1, N_COLS + 1):
            ws.cell(row=row, column=col).fill = section_fill
        row += 1
        continue

    fill = PatternFill('solid', fgColor=section_color.get(section, 'FFFFFF'))
    # Columns: 1=Section 2=Ref 3=Value 4=Package 5=Qty 6=OnHand 7=Ordered
    #          8=Received 9=OrderedFrom 10=MfrPN 11=MouserPN 12=DigiKeyPN
    #          13=Unit$ 14=Ext$ 15=Notes
    mfr = MFR_PN.get(ref, '')
    cells = [section, ref, value, package, qty, '', '', '', '', mfr,
             mouser_pn, digikey_pn,
             unit if unit else None, None, notes]
    for col, val in enumerate(cells, start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        if col in (5, 6, 7, 8, 9, 13, 14):
            c.alignment = center
        if col in (6, 7, 8, 9):
            c.fill = check_fill
        else:
            c.fill = fill
    if qty and qty > 0 and unit:
        ws.cell(row=row, column=14).value = f'=E{row}*M{row}'
        ws.cell(row=row, column=14).number_format = '"$"#,##0.00'
    if unit:
        ws.cell(row=row, column=13).number_format = '"$"#,##0.00'
    row += 1

total_row = row + 1
ws.cell(row=total_row, column=12, value='TOTAL:').font = Font(bold=True)
ws.cell(row=total_row, column=12).alignment = Alignment(horizontal='right')
tc = ws.cell(row=total_row, column=14, value=f'=SUM(N2:N{row-1})')
tc.font = Font(bold=True, color='FFFFFF')
tc.number_format = '"$"#,##0.00'
tc.fill = PatternFill('solid', fgColor='305496')

wsi = wb.create_sheet("Instructions")
instr_lines = [
    'VFO + Keyer Parts Order - Instructions',
    '',
    'Columns:',
    '  Section      VFO / KEYER / POWER / CONN / TEST',
    '  Ref          Reference designator from the KiCad schematic',
    '  Value        Spec to order (value + tolerance + voltage)',
    '  Package      Physical form factor for the PCB footprint',
    '  Qty          Number to order (includes spares where critical)',
    '  On Hand       Mark X if you already have it in stock',
    '  Ordered       Mark X when the PO goes out',
    '  Received      Mark X when the parts arrive',
    '  Ordered From  Which supplier you actually ordered from (M / DK / A / KP)',
    '  Mfr P/N       Manufacturer part number (universal across distributors;',
    '                populated only for active devices and named caps where it is',
    '                unambiguous; blank for generic passives - search by description)',
    '  Mouser P/N    Mouser catalog number (verify before ordering)',
    '  Digi-Key P/N  Digi-Key catalog number (verify before ordering)',
    '  Unit $        Estimated unit price (verify when ordering)',
    '  Ext $         Auto = Qty x Unit $',
    '  Notes         Substitutions, matching requirements, gotchas',
    '',
    'Sections:',
    '  VFO     Si5351 module + jumper hardware + pi pad + 7-pole LPF + J310 buffer',
    '  KEYER   MC1496 + bias network + PNP null injection + LM7171 post-amps + MCP4921 envelope DAC',
    '  POWER   Power input terminal block + bulk bypass + -8.3V derivation from -90V',
    '  CONN    Off-board headers (SPI, RF out, paddle)',
    '  TEST    Bench bring-up extras (BNC tap-offs for scope, etc.)',
    '',
    'Bench bring-up notes:',
    '  - Bench supplies provide +12V, -8.3V, +5V (and -90V if testing on-board zener+dropper)',
    '  - First bring-up: bypass J310 source follower with 0R jumper at Q1 footprint',
    '  - Match PNP transistors T1/T2 by DMM Vbe (~5mV tolerance) for cleanest carrier null',
    '  - Match RLa/RLb (3.9k) by DMM for differential balance (~1% tolerance)',
    '  - MC1496 R4 (51 ohm at pin 1) forms divider with R_F_D (1.5k); sets envelope DAC scale',
    '',
    'NOT included on this list (separate orders):',
    '  - WH2004A LCD display (separate parts list when display section is built)',
    '  - Rotary encoder + paddle key (your choice; not blocking bench bring-up)',
    '  - MCU board (Adafruit Metro ESP32-S3 - already owned)',
    '  - PA-side parts: bias DAC MCP4728, OPA454, cathode monitor, MCP23017 GPIO expander',
    '  - Power supply transformers, mains hardware, chassis',
]
for i, line in enumerate(instr_lines, start=1):
    c = wsi.cell(row=i, column=1, value=line)
    if i == 1:
        c.font = Font(bold=True, size=14)
    elif line.endswith(':') and not line.startswith('  '):
        c.font = Font(bold=True, size=11)
wsi.column_dimensions['A'].width = 110

wb.save('Documentation/VFO_Keyer_Order.xlsx')
print('Saved Documentation/VFO_Keyer_Order.xlsx')
print(f'Total parts rows: {row - 2}')

"""Insert a new 'Mfr P/N' column into VFO_Keyer_Order.xlsx between
'Ordered From' and 'Mouser P/N', populate it for parts where the
manufacturer number is reliably known, and preserve every other cell
(including user-filled check-offs, prices, P/N corrections, notes)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = 'Documentation/VFO_Keyer_Order.xlsx'
INSERT_AT = 10  # before old column 10 (Mouser P/N), after column 9 (Ordered From)

# Mfr P/N lookup keyed on the Ref designator. Conservative - only parts
# where the manufacturer number is unambiguous.
MFR_PN = {
    'U1':    'Adafruit 5566',                  # Adafruit Si5351A STEMMA QT
    'U2':    'MC1496P',                         # ON Semi / Motorola classic DIP-14
    'Q1':    'J310',                            # multi-vendor JFET
    'T1':    '2N3906',                          # multi-vendor PNP
    'U3,4':  'LM7171AIN/NOPB',                  # TI high-speed op-amp DIP-8
    'U5':    'MCP4921-E/P',                     # Microchip DIP-8 industrial
    'D1':    '1N4738A',                         # multi-vendor 8.2 V zener
    'CF1':   'CD15FD221JO3F',                   # Cornell Dubilier silver mica 220 pF 500 V
    'CF3':   'CD15FD391JO3F',                   # Cornell Dubilier 390 pF 500 V
    'CF5':   'CD15FD391JO3F',
    'CF7':   'CD15FD221JO3F',
    'C1':    'CD15FD331JO3F (silver mica option)',
    'CB_BLK':'Kemet T350F106K025AT',            # 10 uF / 25 V tantalum
    'C_F':   'Wima MKP4 680 nF / 63 V',         # film cap
    'XU2':   'Mill-Max 110-43-314 (machined)',
    'XU3,4': 'Mill-Max 110-43-308 (machined)',
}

# Section fill colors (must match the original generator)
SECTION_COLOR = {
    'VFO': 'D9E1F2', 'KEYER': 'E2EFDA', 'POWER': 'FFF2CC',
    'CONN': 'FCE4D6', 'TEST': 'EDEDED',
}
SECTION_FILL = 'D9E1F2'   # for section header rows
CHECK_FILL = 'FFF2CC'
HEADER_FILL = '305496'

thin = Side(border_style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')

wb = openpyxl.load_workbook(XLSX)
ws = wb['VFO + Keyer Order']

# 1. Insert the new column
ws.insert_cols(INSERT_AT)

# 2. Set the column width and header
ws.column_dimensions[get_column_letter(INSERT_AT)].width = 26

hc = ws.cell(row=1, column=INSERT_AT, value='Mfr P/N')
hc.font = Font(bold=True, color='FFFFFF', size=11)
hc.fill = PatternFill('solid', fgColor=HEADER_FILL)
hc.alignment = center
hc.border = border

# 3. For each row, populate Mfr P/N if known, AND apply correct styling
populated = 0
for row_idx in range(2, ws.max_row + 1):
    section_val = ws.cell(row=row_idx, column=1).value
    ref_val = ws.cell(row=row_idx, column=2).value
    desc_val = ws.cell(row=row_idx, column=3).value

    cell = ws.cell(row=row_idx, column=INSERT_AT)
    cell.border = border
    cell.alignment = center

    # Detect section header rows: section in col A, Ref empty,
    # description starts with "---"
    is_section_header = (
        section_val and not ref_val and
        isinstance(desc_val, str) and desc_val.startswith('---')
    )

    if is_section_header:
        cell.fill = PatternFill('solid', fgColor=SECTION_FILL)
    else:
        fill_color = SECTION_COLOR.get(section_val, 'FFFFFF')
        cell.fill = PatternFill('solid', fgColor=fill_color)

        # Look up Mfr P/N for this row
        if ref_val and ref_val in MFR_PN:
            cell.value = MFR_PN[ref_val]
            populated += 1

# 4. Fix the Ext $ formulas (now in column 14, was column 13)
#    Formula was =E{r}*L{r}; needs to become =E{r}*M{r} because Unit$
#    column moved from L (12) to M (13).
ext_col = INSERT_AT + 4   # col 10 + 4 = 14 (Ext $)
unit_col = INSERT_AT + 3  # col 10 + 3 = 13 (Unit $)
formulas_fixed = 0
for row_idx in range(2, ws.max_row + 1):
    cell = ws.cell(row=row_idx, column=ext_col)
    v = cell.value
    if isinstance(v, str) and v.startswith('='):
        # Replace any *L<digits> with *M<digits>
        import re
        new_v = re.sub(r'\*L(\d+)', r'*M\1', v)
        if new_v != v:
            cell.value = new_v
            formulas_fixed += 1

# 5. Fix the SUM total formula (was =SUM(M...:M...), needs =SUM(N...:N...))
sum_fixed = 0
for row_idx in range(2, ws.max_row + 1):
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        v = cell.value
        if isinstance(v, str) and v.startswith('=SUM(M'):
            cell.value = v.replace('SUM(M', 'SUM(N')
            sum_fixed += 1

wb.save(XLSX)
print(f'Inserted column at position {INSERT_AT} (Mfr P/N)')
print(f'Populated {populated} cells with manufacturer part numbers')
print(f'Updated {formulas_fixed} Ext $ row formulas (L -> M)')
print(f'Updated {sum_fixed} SUM total formula(s) (M -> N)')
